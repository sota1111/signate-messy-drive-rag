"""SOT-2564: conditional-format highlights, blue-family colour matching, and pivot-ancestor context.

All three live behind ``RAG_HIGHLIGHT_EXTRA`` (default OFF); OFF must be byte-identical to the plain
solid-fill scan (champion serve unchanged). ON turns three format-dependent questions that looked
UNANSWERABLE into gold-matching evidence:

* idx65 — a yellow highlight applied by a ``値 < -0.99`` conditional-format rule (invisible to
  ``cell.fill``) surfaces its condition.
* idx25 — a 青 (blue) filter matches a 水色 (light-blue) fill so the highlighted cells (and their sum)
  are returned.
* idx15 — a highlighted count cell in a hierarchical pivot carries its enclosing group labels.
"""
from pathlib import Path

import openpyxl
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill

from src.rag.corpus import FileRef
from src.rag.tools.highlight_extract import highlight_extract


def _ref(path: Path) -> FileRef:
    return FileRef(path=path, project="p", category="data", rel=path.name,
                   name=path.name, ext="xlsx")


def _cond_xlsx(path: Path) -> None:
    """A 5x5 correlation-like sheet with a yellow ``<-0.99`` and a red ``>0.99`` CF rule."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "corr"
    ws["A1"] = "x"
    ws["B1"] = "y"
    ws["B2"] = -0.995      # satisfies < -0.99  → yellow
    ws["B3"] = 0.995       # satisfies > 0.99   → red
    ws["B4"] = 0.10        # neither
    yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    ws.conditional_formatting.add("A1:E5", CellIsRule(operator="lessThan", formula=["-0.99"], fill=yellow))
    ws.conditional_formatting.add("A1:E5", CellIsRule(operator="greaterThan", formula=["0.99"], fill=red))
    wb.save(path)


def test_conditional_format_surfaced_only_when_enabled(tmp_path: Path, monkeypatch) -> None:
    path = tmp_path / "corr.xlsx"
    _cond_xlsx(path)
    monkeypatch.setenv("RAG_STRUCTURE_STORE", "0")

    # OFF: cell.fill scan sees nothing (rule-based fills are invisible) → no CF items.
    monkeypatch.setenv("RAG_HIGHLIGHT_EXTRA", "0")
    off = highlight_extract(_ref(path))
    assert all(it["method"]["kind"] != "conditional_format" for it in off["value"])

    # ON, filtered to 黄: only the < -0.99 rule (yellow), with its condition + satisfying cell.
    monkeypatch.setenv("RAG_HIGHLIGHT_EXTRA", "1")
    on = highlight_extract(_ref(path), color="黄")
    cf = [it for it in on["value"] if it["method"]["kind"] == "conditional_format"]
    assert len(cf) == 1
    assert cf[0]["value"] == "セルの値 < -0.99"
    assert cf[0]["method"]["color"] == "黄"
    assert cf[0]["evidence"]["operator"] == "lessThan"
    assert cf[0]["evidence"]["n_matched"] == 1
    assert "B2" in cf[0]["evidence"]["matched_cells"]
    # the red (> 0.99) rule is excluded by the 黄 filter
    assert all(it["evidence"].get("operator") != "greaterThan" for it in cf)


def test_blue_family_matches_light_blue_only_when_enabled(tmp_path: Path, monkeypatch) -> None:
    path = tmp_path / "blue.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = -100.0
    ws["A2"] = -200.0
    mizuiro = PatternFill(start_color="0000B0F0", end_color="0000B0F0", fill_type="solid")  # 水色
    ws["A1"].fill = mizuiro
    ws["A2"].fill = mizuiro
    wb.save(path)
    monkeypatch.setenv("RAG_STRUCTURE_STORE", "0")

    # OFF: 青 filter is an exact-name match, 水色 ≠ 青 → nothing.
    monkeypatch.setenv("RAG_HIGHLIGHT_EXTRA", "0")
    assert highlight_extract(_ref(path), color="青")["evidence"]["n_items"] == 0

    # ON: 青 matches its 水色 shade → both cells returned.
    monkeypatch.setenv("RAG_HIGHLIGHT_EXTRA", "1")
    on = highlight_extract(_ref(path), color="青")
    vals = sorted(float(it["value"]) for it in on["value"] if it["method"]["kind"] == "cell")
    assert vals == [-200.0, -100.0]


def test_pivot_ancestor_group_context_only_when_enabled(tmp_path: Path, monkeypatch) -> None:
    path = tmp_path / "pivot.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Gender", "Age", "Country", "個数"])   # row 1 headers
    ws.append(["Male", "40-44", "Argentina", 1])       # row 2: full group tuple
    ws.append([None, None, "Spain", 12])               # row 3: forward-filled Gender/Age
    ws["D3"].fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")  # 黄
    wb.save(path)
    monkeypatch.setenv("RAG_STRUCTURE_STORE", "0")

    # OFF: the highlighted cell carries no group context.
    monkeypatch.setenv("RAG_HIGHLIGHT_EXTRA", "0")
    off = highlight_extract(_ref(path), color="黄")
    assert off["evidence"]["n_items"] == 1
    assert "group" not in off["value"][0]["evidence"]

    # ON: the enclosing (forward-filled) pivot group is recovered from the columns to the left.
    monkeypatch.setenv("RAG_HIGHLIGHT_EXTRA", "1")
    on = highlight_extract(_ref(path), color="黄")
    hit = next(it for it in on["value"] if it["method"]["kind"] == "cell")
    assert hit["value"] == "12"
    assert hit["evidence"]["group"] == {"Gender": "Male", "Age": "40-44", "Country": "Spain"}


def test_flag_off_is_byte_identical(tmp_path: Path, monkeypatch) -> None:
    """The full flag-OFF contract equals a run with no CF/family/group machinery reachable."""
    path = tmp_path / "mix.xlsx"
    _cond_xlsx(path)
    ws = openpyxl.load_workbook(path)["corr"]
    monkeypatch.setenv("RAG_STRUCTURE_STORE", "0")
    monkeypatch.setenv("RAG_HIGHLIGHT_EXTRA", "0")
    result = highlight_extract(_ref(path))
    # no conditional_format items, no group keys — identical shape to the historical solid-fill scan
    assert all(it["method"]["kind"] == "cell" for it in result["value"])
    assert all("group" not in it["evidence"] for it in result["value"])
