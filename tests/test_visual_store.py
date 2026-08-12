"""SOT-2653 — xlsx 可視事実ストア + serve-path 決定論レーンの offline テスト.

ネットワーク/LLM/実コーパス非依存。in-memory openpyxl ワークブックで抽出器（色分類 / 相関判定 / CF /
行・列ハイライト）を、合成ストア（``visual_store.load`` を monkeypatch）で 4 本の決定論レーン
（idx39 chart / idx65 相関CF / idx82 淡橙行ID / idx97 交差差）と精度優先の deferral、RAG_VISUAL_STORE
既定 OFF の byte-identical 挙動、ツール contract を検証する。
"""
from __future__ import annotations

import io

import openpyxl
import pytest
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill

from src.rag.agent import fact_layer as fl
from src.rag.agent import visual_lane as vl
from src.rag.index import visual_store as vs
from src.rag.tools import contract as _contract


# =========================================================================== colour classification
def _fill(rgb: str) -> PatternFill:
    return PatternFill(patternType="solid", fgColor=rgb)


def test_classify_fill_standard_and_pale_tint():
    assert vs.classify_fill(_fill("FFFFFF00").fgColor) == "黄"          # solid yellow
    # Pale seashell orange (FFF0E6, sat ~10%) — champion classifier drops it, the store recovers オレンジ.
    assert vs.classify_fill(_fill("FFFFF0E6").fgColor) == "オレンジ"
    # Greys / near-white banding stay uncoloured (no false highlight).
    assert vs.classify_fill(_fill("FFF5F5F5").fgColor) is None
    assert vs.classify_fill(_fill("FF4A4A4A").fgColor) is None


def test_is_correlation_bounds():
    assert vs._is_correlation([1.0, -0.5, 0.3, -0.99, 0.8, 1.0, -0.2, 0.4]) is True
    assert vs._is_correlation([1.0, 2.0, -3.0]) is False           # out of [-1,1]
    assert vs._is_correlation([0.1, 0.2]) is False                 # too few cells


# =========================================================================== extractor (in-memory)
def _sheet_record(wb) -> dict:
    return {ws.title: vs._extract_sheet(ws) for ws in wb.worksheets}


def test_extract_static_and_line_highlights():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"], ws["B1"], ws["C1"] = "id", "x", "y"
    for r in range(2, 8):
        ws.cell(r, 1, f"T{r:02d}")
        ws.cell(r, 2, r)
        ws.cell(r, 3, r * 10)
    # A full-column highlight (col B, rows 2..7) then a full-row highlight (row 3, all 3 cols): the row
    # fill wins on B3, leaving col B with 5 yellow cells (≥ threshold) and row 3 fully orange.
    for r in range(2, 8):
        ws.cell(r, 2).fill = _fill("FFFFFF00")           # yellow column
    for c in range(1, 4):
        ws.cell(3, c).fill = _fill("FFFFF0E6")           # pale orange row
    rec = vs._extract_sheet(ws)
    row_hi = {(r["row"], r["color"]) for r in rec["row_highlights"]}
    col_hi = {(c["col_letter"], c["color"]) for c in rec["col_highlights"]}
    assert (3, "オレンジ") in row_hi
    assert ("B", "黄") in col_hi
    # row 3 stores its per-column cells (header→value) so an ID-column lookup is possible.
    r3 = next(r for r in rec["row_highlights"] if r["row"] == 3)
    assert r3["cells"][0]["header"] == "id" and r3["cells"][0]["value"] == "T03"


def test_extract_conditional_format():
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in range(1, 5):
        ws.cell(r, 3, -0.995)
    dxf_fill = PatternFill(bgColor="FFFF00")
    ws.conditional_formatting.add(
        "C1:C4", CellIsRule(operator="lessThan", formula=["-0.99"], fill=dxf_fill))
    rec = vs._extract_sheet(ws)
    assert len(rec["cf"]) == 1
    cf = rec["cf"][0]
    assert cf["operator"] == "lessThan" and cf["formula"] == ["-0.99"]
    assert cf["condition"] == "セルの値 < -0.99"


def test_chart_defined_name_resolution_and_range_parse():
    assert vs._range_column("train!$M$2:$M$8646") == ("train", "M", 2)
    assert vs._range_column("'My Sheet'!$B$3") == ("My Sheet", "B", 3)
    assert vs._range_column("nope") is None


# =========================================================================== synthetic store + lanes
def _records() -> list[dict]:
    aoshio = "プロジェクト/株式会社青潮モビリティサービス/03.データ/train.xlsx"
    shira,ho = "白峰", None
    shiramine = "プロジェクト/白峰信用リスク評価株式会社/03.データ/train.xlsx"
    soshu = "プロジェクト/医療法人社団 蒼泉会 ひがし丘総合病院/02.計画/スケジュール.xlsx"
    aoba = "プロジェクト/株式会社青葉バイオメディカル機器/03.データ/train.xlsx"
    return [
        {"doc_id": aoshio, "project": "株式会社青潮モビリティサービス", "doc_name": "train.xlsx",
         "sheets": {}, "charts": [
             {"member": "chartEx1.xml", "title": "グラフ1",
              "columns": [{"ref": "train!$M$2:$M$8646", "column": "M", "header": "hum", "data_sheet": "train"}]},
             {"member": "chartEx2.xml", "title": "グラフ2",
              "columns": [{"ref": "train!$L$2:$L$8646", "column": "L", "header": "atemp", "data_sheet": "train"}]}]},
        {"doc_id": shiramine, "project": "白峰信用リスク評価株式会社", "doc_name": "train.xlsx",
         "charts": [], "sheets": {
             "Sheet3": {"state": "visible", "is_correlation": True, "static": [], "row_highlights": [],
                        "col_highlights": [], "color_summary": {},
                        "cf": [{"color": "黄", "operator": "lessThan", "formula": ["-0.99"],
                                "condition": "セルの値 < -0.99", "sqref": "A1:Z52"},
                               {"color": "赤", "operator": "greaterThan", "formula": ["0.99"],
                                "condition": "セルの値 > 0.99", "sqref": "A1:Z52"}]},
             # a HIDDEN correlation sheet with a different yellow threshold — must never be chosen.
             "Sheet2": {"state": "hidden", "is_correlation": True, "static": [], "row_highlights": [],
                        "col_highlights": [], "color_summary": {},
                        "cf": [{"color": "黄", "operator": "lessThan", "formula": ["-0.9"],
                                "condition": "セルの値 < -0.9", "sqref": "A1:Z52"}]},
             # a visible NON-correlation sheet (values > 1) with a yellow rule — must be ignored.
             "Sheet1": {"state": "visible", "is_correlation": False, "static": [], "row_highlights": [],
                        "col_highlights": [], "color_summary": {},
                        "cf": [{"color": "黄", "operator": "lessThan", "formula": ["-10000"],
                                "condition": "セルの値 < -10000", "sqref": "A1:Z52"}]}}},
        {"doc_id": soshu, "project": "医療法人社団 蒼泉会 ひがし丘総合病院", "doc_name": "スケジュール.xlsx",
         "charts": [], "sheets": {
             "WBS": {"state": "visible", "is_correlation": False, "cf": [], "col_highlights": [],
                     "color_summary": {"オレンジ": {"n": 60}},
                     "static": [], "row_highlights": [
                         {"row": r, "color": "オレンジ", "n": 12,
                          "cells": [{"col_letter": "A", "header": "No", "value": str(no)},
                                    {"col_letter": "B", "header": "フェーズ", "value": "P"},
                                    {"col_letter": "C", "header": "タスクID", "value": tid}]}
                         for r, no, tid in [(3, 2, "T02"), (15, 14, "T14"), (17, 16, "T16"),
                                            (23, 22, "T22"), (25, 24, "T24")]]}}},
        {"doc_id": aoba, "project": "株式会社青葉バイオメディカル機器", "doc_name": "train.xlsx",
         "charts": [], "sheets": {
             "train": {"state": "visible", "is_correlation": False, "cf": [],
                       "color_summary": {"黄": {"n": 800, "min": 0.0, "max": 19999.0, "range": 19999.0}},
                       "col_highlights": [{"col": 18, "col_letter": "R", "header": "MonthlyIncome",
                                           "color": "黄", "n": 736}],
                       "row_highlights": [{"row": 79, "color": "黄", "n": 33, "cells": []},
                                          {"row": 138, "color": "黄", "n": 33, "cells": []}],
                       "static": [{"cell": "R79", "row": 79, "column": 18, "col_letter": "R",
                                   "value": "10096", "color": "黄", "col_header": "MonthlyIncome"},
                                  {"cell": "R138", "row": 138, "column": 18, "col_letter": "R",
                                   "value": "10368", "color": "黄", "col_header": "MonthlyIncome"}]}}},
    ]


@pytest.fixture
def store(monkeypatch):
    monkeypatch.setattr(vs, "load", lambda path=None: _records())
    monkeypatch.setenv("RAG_VISUAL_STORE", "1")
    monkeypatch.setenv("RAG_FACT_LAYER", "1")
    return None


# --------------------------------------------------------------------------- OFF byte-identical
def test_off_is_inert(monkeypatch):
    monkeypatch.delenv("RAG_VISUAL_STORE", raising=False)
    assert vl.enabled() is False
    assert vl.resolve("青潮モビリティサービスのグラフ1はどのカラムを可視化したものですか。") is None
    assert vl.tool() is None


def test_off_lane_not_in_fact_layer(monkeypatch):
    monkeypatch.setattr(vs, "load", lambda path=None: _records())
    monkeypatch.setenv("RAG_FACT_LAYER", "1")
    monkeypatch.delenv("RAG_VISUAL_STORE", raising=False)
    assert fl.resolve("青潮モビリティサービスのグラフ1はどのカラムを可視化したものですか。", "chart_read") is None
    assert vl.VISUAL_LOOKUP not in {t[0] for t in fl.tools()}


# --------------------------------------------------------------------------- (idx39) chart column
def test_chart_column_lane(store):
    r = vl.resolve("青潮モビリティサービスのtrain.xlsxのSheet1にあるグラフ1はどのカラムを可視化したものですか。")
    assert r is not None and r["value"] == "hum"
    assert r["method"]["selection"] == "chart_value_column"


def test_chart_column_via_fact_layer(store):
    r = fl.resolve("青潮モビリティサービスのグラフ2はどのカラムを可視化したものですか。", "chart_read")
    assert r is not None and r["value"] == "atemp"


def test_chart_defers_unknown_graph_number(store):
    # グラフ9 は存在しない ⇒ 一意束縛できず defer。
    assert vl.resolve("青潮モビリティサービスのグラフ9はどのカラムを可視化したものですか。") is None


# --------------------------------------------------------------------------- (idx65) correlation cond
def test_correlation_condition_lane(store):
    r = vl.resolve("白峰信用リスク評価のtrain.xlsxで、相関係数シートで黄色ハイライトのセルの条件は？")
    assert r is not None and r["value"] == "相関係数が-0.99未満（-0.99より小さい）のセル"
    assert r["evidence"]["sheet"] == "Sheet3"          # visible correlation sheet, not the hidden -0.9 one


def test_correlation_defers_without_color(store):
    assert vl.resolve("白峰信用リスク評価の相関係数シートのハイライトの条件は？") is None


# --------------------------------------------------------------------------- (idx82) highlighted-row IDs
def test_highlight_row_ids_lane(store):
    r = vl.resolve("蒼泉会 ひがし丘総合病院のスケジュール.xlsxのWBSシートで"
                   "オレンジ色にハイライトされている行のタスクIDをすべて教えてください。")
    assert r is not None and r["value"] == "T02、T14、T16、T22、T24"
    assert r["method"]["selection"] == "highlighted_row_ids"


def test_highlight_row_ids_defers_wrong_color(store):
    # 黄色を尋ねても黄ハイライト行は無い ⇒ defer。
    assert vl.resolve("蒼泉会 ひがし丘総合病院のWBSシートで黄色にハイライトされた行のタスクIDをすべて？") is None


# --------------------------------------------------------------------------- (idx97) intersect diff
def test_intersect_diff_lane(store):
    r = vl.resolve("青葉バイオメディカル機器のtrain.xlsxで、"
                   "黄色ハイライトが交差している2つのセルの値の差の絶対値を計算してください。")
    assert r is not None and r["value"] == 272
    assert r["method"]["selection"] == "highlight_intersection_diff"


def test_intersect_ignores_company_name_blue(store):
    # 「青葉」の 青 を色フィルタに拾わない（company-name guard）。黄で 2 交差 → 272。
    assert vl._wanted_color("青葉バイオメディカル機器の黄色ハイライトが交差…") == "黄"


# --------------------------------------------------------------------------- tool contract
def test_tool_lookup(store):
    t = vl.tool()
    assert t is not None and t[0] == vl.VISUAL_LOOKUP
    out = t[3](project="青潮モビリティサービス", kind="chart")
    assert _contract.is_contract(out)
    titles = {c["title"] for d in out["value"] for c in d.get("charts", [])}
    assert "グラフ1" in titles


def test_tool_unknown_project(store):
    out = vl.tool()[3](project="存在しない会社")
    assert out["value"] is None


# --------------------------------------------------------------------------- store schema roundtrip
def test_store_load_schema_roundtrip(tmp_path):
    p = tmp_path / "visual_store.jsonl"
    vs.write_store(_records(), path=p)
    vs.reset_cache()
    rows = vs.load(path=p)
    assert len(rows) == 4
    docs = vs.docs_for_project("青葉バイオメディカル機器", path=p)
    assert len(docs) == 1 and docs[0]["doc_name"] == "train.xlsx"
