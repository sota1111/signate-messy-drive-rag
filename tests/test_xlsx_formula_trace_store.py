"""SOT-2686 — xlsx 数式トレース/回帰適用ストア（cycle7 K3, build 側）の offline テスト（LLM 不要）。

合成 workbook で build の網羅計算を固定する: (a) 黄色ハイライト数式セル → 参照データ行の全属性トレース、
(b) 係数表（切片＋列名付き係数）→ index キーの予測値事前計算。既定 OFF。
"""
from __future__ import annotations

from types import SimpleNamespace

import openpyxl
from openpyxl.styles import PatternFill

from src.rag.index import xlsx_formula_trace as S

_YELLOW = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")


def _make_ref(tmp_path):
    """A synthetic case workbook: a data sheet, a highlighted error formula, and a regression table."""
    wb = openpyxl.Workbook()
    # --- data sheet (>= _DATA_SHEET_MIN_ROWS rows so a formula ref into it names a real data row) ---
    data = wb.active
    data.title = "Sheet1"
    headers = ["id", "YEAR BUILT", "YEAR BUILT_fillna", "SALE PRICE"]
    for c, h in enumerate(headers, 1):
        data.cell(1, c, h)
    for r in range(2, 62):  # 60 data rows
        data.cell(r, 1, f"train_{r - 2}")
        data.cell(r, 2, 1800 + r)
        data.cell(r, 3, r)
        data.cell(r, 4, 1000 * r)
    # --- summary sheet with the yellow highlighted error formula referencing data row 30 ---
    summ = wb.create_sheet("Sheet2")
    summ.cell(17, 1, "切片"); summ.cell(17, 2, 100.0)
    cell = summ.cell(22, 2, "=(Sheet1!B30-Sheet2!B17)^2")
    cell.fill = _YELLOW
    # --- regression table + matching train data sheet with an index column ---
    reg = wb.create_sheet("Reg")
    reg.cell(16, 2, "係数")
    reg.cell(17, 1, "BMI"); reg.cell(17, 2, 0.5)
    reg.cell(18, 1, "Age"); reg.cell(18, 2, 0.1)
    reg.cell(19, 1, "切片"); reg.cell(19, 2, -1.0)
    train = wb.create_sheet("train")
    for c, h in enumerate(["index", "BMI", "Age", "Outcome"], 1):
        train.cell(1, c, h)
    train.cell(2, 1, 1770); train.cell(2, 2, 10.0); train.cell(2, 3, 40.0); train.cell(2, 4, 0)
    train.cell(3, 1, 5); train.cell(3, 2, 20.0); train.cell(3, 3, 30.0); train.cell(3, 4, 1)
    path = tmp_path / "train.xlsx"
    wb.save(path)
    return SimpleNamespace(path=path, rel="プロジェクト/合成/03.データ/train.xlsx",
                           project="合成", name="train.xlsx", ext="xlsx")


def test_enabled_default_off(monkeypatch):
    monkeypatch.delenv("RAG_XLSX_FORMULA_TRACE", raising=False)
    assert S.enabled() is False


def test_highlight_formula_traces_referenced_row(tmp_path):
    rec = S.compute_doc(_make_ref(tmp_path))
    assert rec is not None
    hfs = rec["highlight_formulas"]
    assert len(hfs) == 1 and hfs[0]["cell"] == "B22" and hfs[0]["fill"] == "FFFFFF00"
    rows = hfs[0]["referenced_rows"]
    assert len(rows) == 1 and rows[0]["sheet"] == "Sheet1" and rows[0]["row"] == 30
    attrs = rows[0]["attributes"]
    assert attrs["YEAR BUILT"] == 1830 and attrs["id"] == "train_28"


def test_regression_precomputes_prediction_by_index(tmp_path):
    rec = S.compute_doc(_make_ref(tmp_path))
    regs = rec["regressions"]
    assert len(regs) == 1
    rg = regs[0]
    assert rg["data_sheet"] == "train" and rg["index_column"] == "index"
    # index=1770: -1.0 + 0.5*10 + 0.1*40 = 8.0
    assert abs(rg["predictions"]["1770"] - 8.0) < 1e-9
    assert set(rg["coefficients"]) == {"BMI", "Age"}


def test_build_writes_and_loads_roundtrip(tmp_path, monkeypatch):
    ref = _make_ref(tmp_path)
    out = tmp_path / "store.jsonl"
    S.build([ref], out=out, write_report=False)
    S.reset_cache()
    rows = S.load(out)
    assert len(rows) == 1 and rows[0]["project"] == "合成"
    assert S.docs_for_project("合成", path=out)


def test_no_records_when_nothing_traceable(tmp_path):
    wb = openpyxl.Workbook()
    wb.active.cell(1, 1, "hello")
    p = tmp_path / "plain.xlsx"
    wb.save(p)
    ref = SimpleNamespace(path=p, rel="x/plain.xlsx", project="x", name="plain.xlsx", ext="xlsx")
    assert S.compute_doc(ref) is None
