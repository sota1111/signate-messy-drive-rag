"""SOT-2533 — deterministic structure pre-store: build/persist + zero-regression runtime consult.

The store is a cache of the *same* deterministic extractors, so a tool consulting it must return a
byte-identical result to a live read; a miss / stale entry / disabled flag must fall back to live.
All tests are hermetic (tmp files + a tmp store path), no corpus / no LLM.
"""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl
import pytest
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import PatternFill

from src.rag.corpus import FileRef
from src.rag.index import structure_store


def _make_xlsx(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["月", "売上"])
    ws.append(["1月", 100])
    ws.append(["2月", 200])
    ws["B2"].fill = PatternFill(fill_type="solid", fgColor="FFFF00")  # a highlighted cell
    chart = BarChart()
    data = Reference(ws, min_col=2, min_row=1, max_row=3)
    cats = Reference(ws, min_col=1, min_row=2, max_row=3)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "E5")
    wb.save(path)


def _ref(path: Path) -> FileRef:
    return FileRef(path=path, project="p", category="plan", rel=path.name, name=path.name, ext="xlsx")


@pytest.fixture()
def store_path(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> Path:
    """Point the store's default artifact at a tmp file so the tools read it without a real corpus."""
    out = tmp_path / "structure_store.json"
    monkeypatch.setattr(structure_store, "default_out_path", lambda: out)
    structure_store.reset_cache()
    return out


def _enable(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("RAG_STRUCTURE_STORE", "1")
    structure_store.reset_cache()


def _disable(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("RAG_STRUCTURE_STORE", "0")
    structure_store.reset_cache()


# --------------------------------------------------------------------------- build / persist
def test_file_payload_captures_deterministic_structures(tmp_path: Path) -> None:
    p = tmp_path / "chart.xlsx"
    _make_xlsx(p)
    payload = structure_store._file_payload(_ref(p))

    assert payload["content_sha"]
    assert payload.get("highlights"), "the yellow-filled B2 cell should be captured"
    assert any(str(it["value"]) == "100" for it in payload["highlights"])
    assert payload.get("charts"), "the embedded bar chart should be captured"


def test_write_store_has_schema_header_and_is_reproducible(tmp_path: Path) -> None:
    p = tmp_path / "chart.xlsx"
    _make_xlsx(p)
    ref = _ref(p)
    payload = structure_store._file_payload(ref)
    out = tmp_path / "s.json"

    counts = structure_store.write_store({ref.rel: payload}, {}, out)
    first = out.read_bytes()
    structure_store.write_store({ref.rel: payload}, {}, out)

    doc = json.loads(out.read_text(encoding="utf-8"))
    assert doc["schema"] == structure_store.SCHEMA
    assert doc["version"] == structure_store.SCHEMA_VERSION
    assert counts["files"] == 1 and counts["highlights"] == 1 and counts["charts"] == 1
    assert out.read_bytes() == first  # deterministic / reproducible


def test_build_persists_and_is_guarded(tmp_path: Path, store_path: Path,
                                       monkeypatch: pytest.MonkeyPatch) -> None:
    p = tmp_path / "chart.xlsx"
    _make_xlsx(p)
    ref = _ref(p)
    # keep corpus-wide extractors hermetic (no real corpus walk)
    monkeypatch.setattr(structure_store, "_seating_payload", lambda: None)
    monkeypatch.setattr(structure_store, "_version_diff_payload", lambda: [])

    counts = structure_store.build([ref])
    doc = json.loads(store_path.read_text(encoding="utf-8"))
    assert counts["files"] == 1
    assert ref.rel in doc["files"]
    assert doc["files"][ref.rel]["content_sha"] == structure_store.content_sha(p)


# --------------------------------------------------------------------------- runtime consult
def test_highlight_extract_consult_is_byte_identical(tmp_path: Path, store_path: Path,
                                                     monkeypatch: pytest.MonkeyPatch) -> None:
    from src.rag.tools.highlight_extract import highlight_extract

    p = tmp_path / "h.xlsx"
    _make_xlsx(p)
    ref = _ref(p)

    _disable(monkeypatch)
    live = highlight_extract(ref)
    assert any(str(it["value"]) == "100" for it in live["value"])

    structure_store.write_store({ref.rel: structure_store._file_payload(ref)}, {})
    _enable(monkeypatch)
    assert structure_store.lookup_highlight_items(ref) is not None
    assert highlight_extract(ref) == live  # cached == live (回答不変)


def test_chart_numcache_consult_is_byte_identical(tmp_path: Path, store_path: Path,
                                                  monkeypatch: pytest.MonkeyPatch) -> None:
    from src.rag.tools.chart_numcache import _extract_chart_numcache_live, extract_chart_numcache

    p = tmp_path / "c.xlsx"
    _make_xlsx(p)
    ref = _ref(p)

    live = _extract_chart_numcache_live(ref)
    structure_store.write_store({ref.rel: structure_store._file_payload(ref)}, {})
    _enable(monkeypatch)
    assert extract_chart_numcache(ref) == live


def test_stale_entry_falls_back_to_live(tmp_path: Path, store_path: Path,
                                        monkeypatch: pytest.MonkeyPatch) -> None:
    from src.rag.tools.highlight_extract import highlight_extract

    p = tmp_path / "s.xlsx"
    _make_xlsx(p)
    ref = _ref(p)
    structure_store.write_store({ref.rel: structure_store._file_payload(ref)}, {})

    # Change the source: the stored content_sha no longer matches → the entry is stale.
    wb = openpyxl.load_workbook(p)
    wb.active["B2"].fill = PatternFill(fill_type="solid", fgColor="FF0000")
    wb.save(p)

    _enable(monkeypatch)
    assert structure_store.lookup_highlight_items(ref) is None  # detected as stale

    _disable(monkeypatch)
    live_after = highlight_extract(ref)
    _enable(monkeypatch)
    assert highlight_extract(ref) == live_after  # stale → live re-extraction, not the old cache


def test_disabled_flag_never_consults_store(tmp_path: Path, store_path: Path,
                                            monkeypatch: pytest.MonkeyPatch) -> None:
    from src.rag.tools.highlight_extract import highlight_extract

    p = tmp_path / "d.xlsx"
    _make_xlsx(p)
    ref = _ref(p)
    # A poisoned entry proves the disabled path ignores the store entirely (byte-identical champion).
    structure_store.write_store(
        {ref.rel: {"content_sha": structure_store.content_sha(p),
                   "highlights": [{"value": "POISON", "evidence": {}, "method": {"color": "黄"}}]}},
        {})
    _disable(monkeypatch)

    assert structure_store.lookup_highlight_items(ref) is None
    assert all(it["value"] != "POISON" for it in highlight_extract(ref)["value"])


def test_version_diff_consults_store(tmp_path: Path, store_path: Path,
                                     monkeypatch: pytest.MonkeyPatch) -> None:
    from src.rag.diffpair import Change, VersionPair, structural_diff

    old = tmp_path / "提案書_v1.docx"
    new = tmp_path / "提案書_v2.docx"
    old.write_bytes(b"old")
    new.write_bytes(b"new")
    oref = FileRef(old, "p", "", old.name, old.name, "docx")
    nref = FileRef(new, "p", "", new.name, new.name, "docx")
    pair = VersionPair(oref, nref, "提案書", "rev-suffix")

    structure_store.write_store({}, {"version_diffs": [{
        "old": oref.rel, "new": nref.rel, "base": "提案書", "basis": "rev-suffix",
        "old_sha": structure_store.content_sha(old), "new_sha": structure_store.content_sha(new),
        "changes": [{"label": "QAレビューア", "before": "池田直哉", "after": "小林直樹",
                     "kind": "modify"}]}]})
    _enable(monkeypatch)

    # The dummy .docx files are never parsed: the fresh cached diff short-circuits the live read.
    assert structural_diff(pair) == [Change("QAレビューア", "池田直哉", "小林直樹", "modify")]


def test_seating_consults_store(tmp_path: Path, store_path: Path,
                                monkeypatch: pytest.MonkeyPatch) -> None:
    from src.rag.tools.seating_chart import Seat, build_directory

    f = tmp_path / "座席表.pptx"
    f.write_bytes(b"dummy-pptx")
    ref = FileRef(f, "社内管理", "internal", f.name, f.name, "pptx")
    structure_store.write_store({}, {"seating": {
        "file": ref.rel, "content_sha": structure_store.content_sha(f), "pixel_sha": "abc123",
        "origin": "reviewed",
        "seats": [{"name": "佐藤", "ext": "7002", "role": "PM", "pod": 1, "row": 0, "col": 0}]}})
    _enable(monkeypatch)

    directory = build_directory(ref)  # returns from the store before any image decode
    assert directory.origin == "reviewed"
    assert directory.pixel_sha == "abc123"
    assert directory.seats == (Seat("佐藤", "7002", "PM", 1, 0, 0),)
