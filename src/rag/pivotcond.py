"""PivotTable / AutoFilter condition reading — answer "抽出条件" questions deterministically.

A class of questions asks to read the *extraction condition* baked into an ``.xlsx``:

  * **PivotTable** (valid idx6 かえで train.xlsx: ALPの平均が最も高い層 → Gender=Male、disease=1、Age=68
    に対する 平均 / ALP; idx21 青葉バイオ train.xlsx Pivotシート: 平均月収が最も高い層 →
    Attrition=No、Gender=Female、MaritalStatus=Single、EducationField=Human Resources);
  * **applied AutoFilter** (valid idx11 東都 train.xlsx trainシート: Gender=Male、Country=India、target=2).

Plain retrieval / LLM reading of the flattened sheet fails these because the condition lives in the
workbook's *pivot definition* and *autofilter* XML, not in visible cell text. This module reads those
structures directly and recomputes the answer:

  * PivotTable: parse the pivot definition (row fields + data fields + aggregation) and its cache
    source, then recompute the aggregation over the source table grouped by the full row-field tuple
    and report the leaf group that maximises (or minimises) the asked measure, as
    ``列=値、…（で抽出されたデータに対する集計名）``.
  * AutoFilter: read each active ``filterColumn`` and map its column offset to the header name, as
    ``列=値、…``.

It is *additive-safe* like ``src.rag.diffpair``: only questions that explicitly reference a pivot /
filter extraction condition are routed here, and any file that cannot be resolved to a single
unambiguous answer abstains (returns ``None``) — Missing (0) beats Incorrect (−1) under the rubric.

All heavy deps (openpyxl) and corpus access are imported/read lazily inside functions and guarded, so
importing this module at serve time (lean container, no corpus) is free and never raises.
"""
from __future__ import annotations

import re
import zipfile
from dataclasses import dataclass, field
from statistics import mean

from src.rag.corpus import nfc

_NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"

# ---- question routing (kept in sync with archetype.classify via is_pivot_condition_question) ----
_PIVOT_RE = re.compile(r"pivot|ピボット", re.IGNORECASE)
_FILTER_RE = re.compile(r"フィルタ")  # matches フィルタ / フィルター
_COND_RE = re.compile(r"抽出条件|抽出|条件|集計|最も高い|最も低い|最大|最小|一番")
_MAX_RE = re.compile(r"最も高い|最大|一番高い|最も大きい|一番大きい|highest|max", re.IGNORECASE)
_MIN_RE = re.compile(r"最も低い|最小|一番低い|最も小さい|一番小さい|lowest|min", re.IGNORECASE)
_AGG_ASK_RE = re.compile(r"集計")


def is_pivot_condition_question(question: str) -> bool:
    """True for a PivotTable / AutoFilter extraction-condition question (else leave to retrieval)."""
    q = nfc(question)
    is_pivot = bool(_PIVOT_RE.search(q))
    is_filter = bool(_FILTER_RE.search(q) and re.search(r"抽出|条件", q))
    if not (is_pivot or is_filter):
        return False
    return bool(_COND_RE.search(q))


def _is_filter_question(q: str) -> bool:
    """A filter question references a フィルタ but no pivot (pivot takes precedence when both)."""
    return bool(_FILTER_RE.search(q)) and not _PIVOT_RE.search(q)


# --------------------------------------------------------------------------------------------
def _fmt_val(v) -> str | None:
    if v is None:
        return None
    if isinstance(v, bool):
        return str(v)
    if isinstance(v, float):
        return str(int(v)) if v.is_integer() else str(v)
    if isinstance(v, int):
        return str(v)
    s = nfc(str(v)).strip()
    return s or None


def _col_index(ref_cell: str) -> int:
    """0-based column index of an A1-style cell reference's column part (A→0, B→1, …)."""
    letters = re.match(r"([A-Za-z]+)", ref_cell)
    if not letters:
        return 0
    idx = 0
    for ch in letters.group(1).upper():
        idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return idx - 1


# ---------------------------------------- pivot ---------------------------------------------
@dataclass
class PivotSpec:
    sheet: str                          # cache worksheetSource sheet name
    col_offset: int                     # 0-based column offset of the source range's first column
    fields: list[str]                   # cache field names (index-aligned)
    row_fields: list[int]               # cache field indices used as row fields (in order)
    data_fields: list[tuple[str, int, str]]  # (display name, cache field index, subtotal func)


def _pivot_specs(path) -> list[PivotSpec]:
    """Parse every pivot table in a workbook into a PivotSpec (empty list if none/unreadable)."""
    specs: list[PivotSpec] = []
    try:
        z = zipfile.ZipFile(str(path))
    except Exception:
        return specs
    with z:
        pt_names = sorted(n for n in z.namelist()
                          if re.match(r"xl/pivotTables/pivotTable\d+\.xml$", n))
        for pt_name in pt_names:
            try:
                spec = _parse_one_pivot(z, pt_name)
            except Exception:
                spec = None
            if spec is not None:
                specs.append(spec)
    return specs


def _parse_one_pivot(z: zipfile.ZipFile, pt_name: str) -> PivotSpec | None:
    import xml.etree.ElementTree as ET

    pt = ET.fromstring(z.read(pt_name))
    row_fields = [int(f.get("x")) for f in pt.findall(f"{_NS}rowFields/{_NS}field")
                  if int(f.get("x", "-1")) >= 0]
    data_fields: list[tuple[str, int, str]] = []
    for df in pt.findall(f"{_NS}dataFields/{_NS}dataField"):
        fld = df.get("fld")
        if fld is None:
            continue
        data_fields.append((df.get("name") or "", int(fld), df.get("subtotal") or "average"))
    if not row_fields or not data_fields:
        return None

    # follow the rel to the pivot cache definition to read field names + source range
    base = pt_name.rsplit("/", 1)[-1]
    rels_name = f"xl/pivotTables/_rels/{base}.rels"
    cache_def = None
    try:
        rels = ET.fromstring(z.read(rels_name))
        for r in rels:
            tgt = r.get("Target", "")
            if "pivotCacheDefinition" in tgt:
                cache_def = tgt.split("/")[-1]
                break
    except Exception:
        cache_def = None
    if not cache_def:
        # fall back to the conventional name matching the pivot index
        m = re.search(r"pivotTable(\d+)", pt_name)
        cache_def = f"pivotCacheDefinition{m.group(1) if m else '1'}.xml"
    cdef = ET.fromstring(z.read(f"xl/pivotCache/{cache_def}"))
    fields = [cf.get("name") or "" for cf in cdef.findall(f"{_NS}cacheFields/{_NS}cacheField")]
    ws_src = cdef.find(f"{_NS}cacheSource/{_NS}worksheetSource")
    if ws_src is None or ws_src.get("sheet") is None:
        return None
    ref = ws_src.get("ref") or "A1"
    col_offset = _col_index(ref.split(":")[0])
    return PivotSpec(sheet=nfc(ws_src.get("sheet")), col_offset=col_offset, fields=fields,
                     row_fields=row_fields, data_fields=data_fields)


def _agg(values: list, subtotal: str):
    nums = [float(v) for v in values
            if isinstance(v, (int, float)) and not isinstance(v, bool)]
    s = subtotal.lower()
    if s in ("sum",):
        return sum(nums) if nums else None
    if s == "count":
        return sum(v is not None for v in values)
    if s == "countnums":
        return len(nums)
    if s in ("max",):
        return max(nums) if nums else None
    if s in ("min",):
        return min(nums) if nums else None
    # default / average
    return mean(nums) if nums else None


def _source_rows(path, sheet: str) -> list[tuple]:
    import openpyxl

    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    try:
        if sheet not in wb.sheetnames:
            return []
        ws = wb[sheet]
        return list(ws.iter_rows(values_only=True))
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _pick_data_field(spec: PivotSpec, question: str) -> tuple[str, int, str] | None:
    """Choose which pivot measure the question asks about.

    Match a data field whose *underlying cache field name* appears in the question (longest match
    wins). If exactly one measure exists, use it unconditionally. Otherwise abstain (None)."""
    if len(spec.data_fields) == 1:
        return spec.data_fields[0]
    q = nfc(question).lower()
    best, best_len = None, 0
    for name, fld, sub in spec.data_fields:
        base = spec.fields[fld].lower() if 0 <= fld < len(spec.fields) else ""
        if base and base in q and len(base) > best_len:
            best, best_len = (name, fld, sub), len(base)
    return best


def _compute_pivot(path, spec: PivotSpec, question: str) -> str | None:
    df = _pick_data_field(spec, question)
    if df is None:
        return None
    dfname, dfld, subtotal = df
    q = nfc(question)
    if _MIN_RE.search(q):
        direction = "min"
    elif _MAX_RE.search(q):
        direction = "max"
    else:
        return None  # no superlative → cannot pick a unique leaf group

    rows = _source_rows(path, spec.sheet)
    if len(rows) < 2:
        return None
    data = rows[1:]  # row 0 = header

    from collections import defaultdict
    groups: dict[tuple, list] = defaultdict(list)
    for r in data:
        key = tuple(r[spec.col_offset + i] if spec.col_offset + i < len(r) else None
                    for i in spec.row_fields)
        if any(k is None for k in key):  # only fully-specified leaf groups form a condition
            continue
        val = r[spec.col_offset + dfld] if spec.col_offset + dfld < len(r) else None
        groups[key].append(val)

    agg = {k: _agg(v, subtotal) for k, v in groups.items()}
    agg = {k: v for k, v in agg.items() if v is not None}
    if not agg:
        return None
    best = (max if direction == "max" else min)(agg, key=agg.get)

    conds = []
    for pos, fidx in enumerate(spec.row_fields):
        name = spec.fields[fidx] if 0 <= fidx < len(spec.fields) else f"col{fidx}"
        val = _fmt_val(best[pos])
        if val is None:
            return None
        conds.append(f"{nfc(name)}={val}")
    cond_str = "、".join(conds)
    if _AGG_ASK_RE.search(q):
        return f"{cond_str}で抽出されたデータに対する{nfc(dfname)}"
    return cond_str


# --------------------------------------- autofilter -----------------------------------------
@dataclass
class FilterSpec:
    sheet: str
    conditions: list[str] = field(default_factory=list)  # "列=値" rendered


def _filter_specs(path) -> list[FilterSpec]:
    """Active AutoFilter conditions per worksheet (empty when none)."""
    import openpyxl

    out: list[FilterSpec] = []
    try:
        wb = openpyxl.load_workbook(str(path), data_only=True)
    except Exception:
        return out
    for ws in wb.worksheets:
        af = getattr(ws, "auto_filter", None)
        if not af or not af.ref or not getattr(af, "filterColumn", None):
            continue
        top_left = af.ref.split(":")[0]
        start_col = _col_index(top_left)                       # 0-based
        start_row = int(re.search(r"(\d+)", top_left).group(1))
        conds: list[str] = []
        for fc in af.filterColumn:
            col_id = getattr(fc, "colId", None)
            if col_id is None:
                continue
            header = ws.cell(start_row, start_col + int(col_id) + 1).value
            vals = _filter_values(fc)
            if header is None or not vals:
                continue
            conds.append(f"{nfc(str(header)).strip()}={'/'.join(vals)}")
        if conds:
            out.append(FilterSpec(sheet=nfc(ws.title), conditions=conds))
    return out


def _filter_values(fc) -> list[str]:
    vals: list[str] = []
    filters = getattr(fc, "filters", None)
    if filters is not None:
        for f in (getattr(filters, "filter", None) or []):
            v = getattr(f, "val", f)
            s = _fmt_val(v)
            if s is not None:
                vals.append(s)
    customs = getattr(fc, "customFilters", None)
    if customs is not None:
        for cf in (getattr(customs, "customFilter", None) or []):
            v = _fmt_val(getattr(cf, "val", None))
            op = getattr(cf, "operator", None)
            if v is not None:
                op_sym = {"greaterThan": ">", "lessThan": "<", "greaterThanOrEqual": ">=",
                          "lessThanOrEqual": "<=", "notEqual": "≠"}.get(op or "", "")
                vals.append(f"{op_sym}{v}")
    return vals


def _compute_filter(path, question: str) -> str | None:
    specs = _filter_specs(path)
    if not specs:
        return None
    q = nfc(question)
    # prefer a sheet named in the question (e.g. "trainシート"); else the single active-filter sheet
    named = [s for s in specs if s.sheet and s.sheet in q]
    chosen = named[0] if named else (specs[0] if len(specs) == 1 else None)
    if chosen is None:
        return None
    return "、".join(chosen.conditions)


# ------------------------------------ resolution / answer -----------------------------------
def _candidate_files(question: str):
    from src.rag import corpus  # lazy: corpus may be absent at serve time
    from src.rag.extract import glossary

    try:
        refs = corpus.walk()
    except Exception:
        return []
    xlsx = [r for r in refs if r.ext == "xlsx" and not r.name.startswith("~$")]
    try:
        company = glossary.load().company_of(question)
    except Exception:
        company = None
    if company:
        c = nfc(company)
        scoped = [r for r in xlsx if nfc(r.project) and (nfc(r.project) in c or c in nfc(r.project))]
        if scoped:
            xlsx = scoped
    q = nfc(question)
    named = [r for r in xlsx if nfc(r.name) in q]
    return named or xlsx


def answer_question(question: str) -> str | None:
    """Deterministic answer for a pivot / filter condition question, or None to abstain.

    Resolves the referenced workbook (by company + filename), then reads the pivot definition or the
    active autofilter and reports its extraction condition. Returns None whenever the target cannot be
    resolved to a single unambiguous answer, so the caller abstains (Missing 0 beats Incorrect −1)."""
    if not is_pivot_condition_question(question):
        return None
    try:
        cands = _candidate_files(question)
    except Exception:
        return None  # defensive: any resolution failure → abstain
    if not cands:
        return None

    want_filter = _is_filter_question(nfc(question))
    answers: list[str] = []
    for ref in cands:
        try:
            if want_filter:
                a = _compute_filter(ref.path, question)
            else:
                specs = _pivot_specs(ref.path)
                a = None
                for spec in specs:
                    a = _compute_pivot(ref.path, spec, question)
                    if a:
                        break
        except Exception:
            a = None
        if a:
            answers.append(a)
    # unique, non-empty answer only (ambiguous → abstain)
    uniq = list(dict.fromkeys(answers))
    if len(uniq) == 1:
        return uniq[0]
    return None


# ---------------------------------------- benchmark -----------------------------------------
@dataclass
class BenchItem:
    question: str
    truth: str
    company: str
    source: str


def benchmark_items() -> list[BenchItem]:
    """One labelled Q/A per pivot measure and per active autofilter in the corpus.

    Ground truth is this module's own deterministic answer, so scoring a RAG run against it measures
    whether the generator correctly *routes* a condition question here and reproduces the answer — the
    same path that answers valid idx6 / idx11 / idx21."""
    from src.rag import corpus

    try:
        refs = [r for r in corpus.walk() if r.ext == "xlsx" and not r.name.startswith("~$")]
    except Exception:
        return []
    items: list[BenchItem] = []
    for ref in refs:
        company = nfc(ref.project)
        if not company:
            continue
        # pivot measures
        try:
            specs = _pivot_specs(ref.path)
        except Exception:
            specs = []
        for spec in specs:
            for dfname, fld, _sub in spec.data_fields:
                base = spec.fields[fld] if 0 <= fld < len(spec.fields) else ""
                if not base:
                    continue
                q = (f"{company}の{ref.name}のPivotシートにおいて、"
                     f"{base}の平均が最も高い層の抽出条件と集計内容を答えてください。")
                try:
                    truth = answer_question(q)
                except Exception:
                    truth = None
                if truth:
                    items.append(BenchItem(q, truth, company, ref.rel))
        # active autofilter (one item per file)
        try:
            fspecs = _filter_specs(ref.path)
        except Exception:
            fspecs = []
        if fspecs:
            sheet = fspecs[0].sheet
            q = (f"{company}の{ref.name}において、{sheet}シートでフィルターで"
                 "抽出されている条件を教えてください。")
            try:
                truth = answer_question(q)
            except Exception:
                truth = None
            if truth:
                items.append(BenchItem(q, truth, company, ref.rel))
    return items
