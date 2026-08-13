"""xlsx 数式依存トレース＋記載回帰係数の行適用の build-time ストア（SOT-2686 / cycle7 K3）.

Sonnet gold100 cycle7 の abstain のうち idx47 / idx83 は、証拠が「セルのテキスト」ではなく
**xlsx の計算構造**にあり、file_grep / FTS では到達できないのが主因だった
(`docs/ai/sonnet_cycle_analysis/cycle7.md` §1 クラスタ K3)。

* **idx47**（青嶺 train.xlsx）: 黄色ハイライトセル ``B22`` は誤差 ``(予測−実測)^2`` を計算しており、その
  数式が ``Sheet1!U26118`` 等の **データ行 26118 を参照**している。その行の ``YEAR BUILT``（=1899）を辿る
  手段が無かった。⇒ **数式依存トレース**: ハイライト数式セルの参照セル群を解析し、参照先の *データ行* の
  全属性（id ＋ 各列ヘッダ→値）を焼き込む。
* **idx83**（みなみ野 train.xlsx）: 回帰分析シートの **係数表**（切片＋列名付き係数）を、``train`` シートの
  ``index=1770`` 行へ当てはめた予測値（=0.38317）が compute 数手では届かなかった。⇒ **記載回帰係数の行適用**:
  係数表を網羅検出し、係数×全行の予測値を **index をキーに事前計算**して焼き込む。

本モジュールは build 時に一度だけ、xlsx/xlsm を **質問を見ずに全数** 走査する。genai 呼び出しゼロ・
openpyxl のみ（`derived_metrics` / `visual_store` と同一の LLM-free・追加的・fail-open 規律）。
読み出し/配線は :mod:`src.rag.agent.xlsx_formula_lane`。

Design invariants
-----------------
* **Opt-in at serve time.** :func:`enabled` (``RAG_XLSX_FORMULA_TRACE``) が runtime 参照のみ gate する。
  default OFF ⇒ champion serve path は byte-identical。
* **Build は LLM フリー・追加的.** openpyxl で数式文字列・セル値・係数表を読むだけ。読めない workbook は
  1 件スキップして継続。
* **Question-independent.** universe は 03.データ/02.計画 等の全 xlsx。質問も gold も参照せず、全ハイライト
  数式・全係数表を網羅計算する。
* **No hardcoding.** gold 値・idx 番号を一切持たない。全値は原ファイルから読み、出典（doc/sheet/cell）付き。
* **Fail-open.** artifact 欠落・解析不能はすべて空へフォールバック（回帰ゼロ）。
"""
from __future__ import annotations

import io
import json
import os
import re
import unicodedata
from pathlib import Path
from typing import Any, Mapping, Sequence

from config import settings
from src.rag.corpus import FileRef, nfc, walk

SCHEMA = "xlsx-formula-trace"
SCHEMA_VERSION = 1

_ON = {"1", "true", "yes", "on"}
_XLSX_EXTS = {"xlsx", "xlsm"}

# A worksheet counts as a "data sheet" (a table of records, not a small summary/coefficient block) when
# it has at least this many rows — so a highlighted formula's reference into it names a genuine data row.
_DATA_SHEET_MIN_ROWS = 50

# Excel A1 reference inside a formula: optional 'Sheet'! prefix, absolute markers tolerated.
_REF_RE = re.compile(r"(?:'([^']+)'|([A-Za-z0-9_぀-鿿]+))?!?\$?([A-Za-z]{1,3})\$?(\d+)")
# The above is greedy about sheet names; we only trust a sheet prefix when a literal '!' precedes the cell.
_SHEETREF_RE = re.compile(r"(?:'([^']+)'!|([A-Za-z0-9_぀-鿿]+)!)?\$?([A-Za-z]{1,3})\$?(\d+)")

# Header text that marks the coefficient column of a regression-summary table.
_COEF_HEADERS = {"係数", "coefficient", "coefficients", "coef", "coefs", "estimate", "回帰係数"}
# Row labels that name the intercept (constant term) rather than a feature column.
_INTERCEPT_LABELS = {"切片", "定数項", "定数", "intercept", "const", "constant", "(intercept)", "_cons"}
# Header names of an index column that "index=N" questions reference.
_INDEX_HEADERS = {"index", "インデックス"}


def enabled() -> bool:
    """True when the serve path may consult the xlsx formula-trace store (default OFF — opt-in)."""
    return os.getenv("RAG_XLSX_FORMULA_TRACE", "0").strip().lower() in _ON


def default_out_path() -> Path:
    return settings.ARTIFACTS_DIR / "xlsx_formula_trace.jsonl"


def default_report_path() -> Path:
    return settings.ARTIFACTS_DIR / "xlsx_formula_trace_build_report.json"


# --------------------------------------------------------------------------- case-name identity key
_CORP = r"(?:株式会社|医療法人社団|一般社団法人|一般財団法人|有限会社|合同会社|合資会社)"
_CORP_PREFIX = re.compile(rf"^{_CORP}\s*")
_CORP_SUFFIX = re.compile(rf"\s*{_CORP}$")


def owner_key(value: Any) -> str:
    """案件名の識別キー（法人格・空白・ケースを吸収）— visual_store / action_row_store と同じ規約。"""
    s = unicodedata.normalize("NFKC", str(value or ""))
    s = _CORP_PREFIX.sub("", s)
    s = _CORP_SUFFIX.sub("", s)
    return re.sub(r"[\s　]", "", s).lower()


def _norm_header(value: Any) -> str:
    return unicodedata.normalize("NFKC", str(value if value is not None else "")).replace(" ", "").strip().lower()


def _is_number(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def _fill_rgb(cell) -> str | None:
    try:
        fill = cell.fill
        if fill and fill.patternType:
            rgb = fill.fgColor.rgb
            if isinstance(rgb, str):
                return rgb
    except Exception:  # noqa: BLE001
        return None
    return None


def _is_highlight(rgb: str | None) -> bool:
    """A deliberately coloured cell (solid fill that is not white/none)."""
    if not rgb or not isinstance(rgb, str):
        return False
    up = rgb.upper()
    return up not in ("00000000", "FFFFFFFF", "FFFFFF", "FF000000")


# --------------------------------------------------------------------------- formula-reference tracing
def _data_sheets(wb) -> dict[str, Any]:
    """``{worksheet.title: worksheet}`` for sheets large enough to hold record rows."""
    return {ws.title: ws for ws in wb.worksheets if ws.max_row >= _DATA_SHEET_MIN_ROWS}


def _row_attributes(ws_vals, row: int) -> dict[str, Any]:
    """``{header(row1) → value(row)}`` for one data row (JSON-safe scalars only)."""
    attrs: dict[str, Any] = {}
    for col in range(1, ws_vals.max_column + 1):
        header = ws_vals.cell(1, col).value
        if header is None or str(header).strip() == "":
            continue
        val = ws_vals.cell(row, col).value
        attrs[nfc(str(header)).strip()] = _jsonable(val)
    return attrs


def _jsonable(val: Any) -> Any:
    if val is None or isinstance(val, (int, float, str, bool)):
        return val
    return str(val)


def _trace_formula_cell(formula: str, own_sheet: str, data_sheets: Mapping[str, Any],
                        vals_by_sheet: Mapping[str, Any]) -> list[dict[str, Any]]:
    """Resolve the *data rows* a highlighted formula references (unique (sheet,row) pairs)."""
    seen: set[tuple[str, int]] = set()
    refs: list[dict[str, Any]] = []
    for m in _SHEETREF_RE.finditer(formula):
        sheet = m.group(1) or m.group(2) or own_sheet
        row = int(m.group(4))
        if sheet not in data_sheets or row <= 1:
            continue
        ws = data_sheets[sheet]
        if row > ws.max_row:
            continue
        key = (sheet, row)
        if key in seen:
            continue
        seen.add(key)
        ws_vals = vals_by_sheet.get(sheet)
        attrs = _row_attributes(ws_vals, row) if ws_vals is not None else {}
        refs.append({
            "sheet": sheet,
            "row": row,
            "id": attrs.get("id"),
            "attributes": attrs,
        })
    return refs


def _highlight_formulas(wb, wb_vals) -> list[dict[str, Any]]:
    """Every solid-highlighted formula cell that references at least one data row, with the traced rows."""
    data_sheets = _data_sheets(wb)
    vals_by_sheet = {ws.title: ws for ws in wb_vals.worksheets}
    out: list[dict[str, Any]] = []
    for ws in wb.worksheets:
        # Skip enormous sheets when scanning for the (small) highlighted formula block — formulas that are
        # highlighted markers live in compact summary sheets, not million-row data tables.
        max_scan = ws.max_row
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if not (isinstance(val, str) and val.startswith("=")):
                    continue
                rgb = _fill_rgb(cell)
                if not _is_highlight(rgb):
                    continue
                traced = _trace_formula_cell(val, ws.title, data_sheets, vals_by_sheet)
                if not traced:
                    continue
                out.append({
                    "sheet": ws.title,
                    "cell": cell.coordinate,
                    "formula": val,
                    "fill": rgb,
                    "referenced_rows": traced,
                })
        _ = max_scan
    return out


# --------------------------------------------------------------------------- documented-regression apply
def _find_coef_tables(wb_vals) -> list[dict[str, Any]]:
    """Detect regression coefficient tables: a '係数' column header with feature/intercept rows below it."""
    tables: list[dict[str, Any]] = []
    for ws in wb_vals.worksheets:
        if ws.max_row > 2000 or ws.max_column > 40:  # coefficient tables live in small summary sheets
            continue
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                if _norm_header(ws.cell(r, c).value) not in _COEF_HEADERS:
                    continue
                coefficients: dict[str, float] = {}
                intercept: float | None = None
                rr = r + 1
                blanks = 0
                while rr <= ws.max_row and blanks < 2:
                    coef = ws.cell(rr, c).value
                    label = None
                    for lc in range(1, c):  # label = leftmost non-empty cell left of the coef column
                        v = ws.cell(rr, lc).value
                        if v is not None and str(v).strip() != "":
                            label = str(v).strip()
                            break
                    if label is None or not _is_number(coef):
                        blanks += 1
                        rr += 1
                        continue
                    blanks = 0
                    if _norm_header(label) in _INTERCEPT_LABELS:
                        intercept = float(coef)
                    else:
                        coefficients[nfc(label)] = float(coef)
                    rr += 1
                if intercept is not None and len(coefficients) >= 2:
                    tables.append({"sheet": ws.title, "coef_cell": ws.cell(r, c).coordinate,
                                   "intercept": intercept, "coefficients": coefficients})
    return tables


def _apply_regression(wb_vals, table: Mapping[str, Any]) -> dict[str, Any] | None:
    """Precompute predictions (intercept + Σ coef·feature) per index value over a matching data sheet."""
    features = table["coefficients"]
    feat_norms = {_norm_header(f): f for f in features}
    for ws in wb_vals.worksheets:
        if ws.max_row < 2:
            continue
        headers = {}
        for col in range(1, ws.max_column + 1):
            h = _norm_header(ws.cell(1, col).value)
            if h:
                headers[h] = col
        index_col = next((headers[h] for h in _INDEX_HEADERS if h in headers), None)
        if index_col is None:
            continue
        if not all(fn in headers for fn in feat_norms):
            continue
        feat_cols = {orig: headers[fn] for fn, orig in feat_norms.items()}
        predictions: dict[str, float] = {}
        for row in range(2, ws.max_row + 1):
            idxval = ws.cell(row, index_col).value
            if idxval is None:
                continue
            try:
                pred = float(table["intercept"])
                ok = True
                for feat, coef in features.items():
                    v = ws.cell(row, feat_cols[feat]).value
                    if not _is_number(v):
                        ok = False
                        break
                    pred += float(coef) * float(v)
                if not ok:
                    continue
            except Exception:  # noqa: BLE001
                continue
            key = str(int(idxval)) if isinstance(idxval, float) and idxval.is_integer() else str(idxval)
            predictions[key] = pred
        if predictions:
            index_header = nfc(str(ws.cell(1, index_col).value)).strip()
            return {
                "coef_sheet": table["sheet"],
                "coef_cell": table["coef_cell"],
                "intercept": table["intercept"],
                "coefficients": features,
                "data_sheet": ws.title,
                "index_column": index_header,
                "predictions": predictions,
            }
    return None


# --------------------------------------------------------------------------- per-document record
def compute_doc(ref: FileRef) -> dict[str, Any] | None:
    """Build one workbook's formula-trace record (``None`` when nothing traceable — no fabrication)."""
    import openpyxl
    data = ref.path.read_bytes()
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=False)
    wb_vals = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    highlight_formulas = _highlight_formulas(wb, wb_vals)
    regressions: list[dict[str, Any]] = []
    for table in _find_coef_tables(wb_vals):
        applied = _apply_regression(wb_vals, table)
        if applied is not None:
            regressions.append(applied)
    if not highlight_formulas and not regressions:
        return None
    return {
        "doc_id": nfc(ref.rel), "project": nfc(ref.project), "doc_name": nfc(ref.name),
        "highlight_formulas": highlight_formulas, "regressions": regressions,
    }


# --------------------------------------------------------------------------- universe / build / io
def _universe(refs: Sequence[FileRef]) -> list[FileRef]:
    out: list[FileRef] = []
    for r in refs:
        if r.ext not in _XLSX_EXTS:
            continue
        name = nfc(r.name).lower()
        if name.startswith("~$") or "old" in name:
            continue
        out.append(r)
    out.sort(key=lambda r: r.rel)
    return out


def write_store(records: Sequence[Mapping[str, Any]], path: Path | None = None) -> dict[str, int]:
    """Atomically write a reproducible JSONL store (schema header + doc_id-sorted rows)."""
    out = Path(path) if path is not None else default_out_path()
    out.parent.mkdir(parents=True, exist_ok=True)
    ordered = sorted(records, key=lambda r: r.get("doc_id", ""))
    tmp = out.with_suffix(out.suffix + ".tmp")
    with tmp.open("w", encoding="utf-8", newline="\n") as handle:
        handle.write(json.dumps({"schema": SCHEMA, "version": SCHEMA_VERSION},
                                ensure_ascii=False, sort_keys=True) + "\n")
        for rec in ordered:
            handle.write(json.dumps(rec, ensure_ascii=False, sort_keys=True) + "\n")
    tmp.replace(out)
    reset_cache()
    return {"docs": len(ordered)}


def build(refs: Sequence[FileRef] | None = None, *, out: Path | None = None,
          write_report: bool = True) -> dict[str, Any]:
    """Scan every xlsx and bake the formula-trace store (question-independent, LLM-free, fail-open)."""
    refs = list(refs) if refs is not None else list(walk())
    universe = _universe(refs)
    records: list[dict[str, Any]] = []
    skipped: list[str] = []
    for ref in universe:
        try:
            rec = compute_doc(ref)
        except Exception:  # noqa: BLE001 — one bad workbook must not sink the build
            rec = None
        if rec is not None:
            records.append(rec)
        else:
            skipped.append(nfc(ref.rel))
    stats = write_store(records, out)
    report = {
        "schema": SCHEMA, "version": SCHEMA_VERSION,
        "universe": len(universe), "records": len(records), "skipped": skipped,
        "highlight_formulas": sum(len(r.get("highlight_formulas", [])) for r in records),
        "regressions": sum(len(r.get("regressions", [])) for r in records),
    }
    if write_report:
        rp = default_report_path()
        rp.parent.mkdir(parents=True, exist_ok=True)
        rp.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"records": len(records), "docs": stats["docs"], "report": report}


# --------------------------------------------------------------------------- load / read (minimal API)
_LOAD_CACHE: dict[str, list[dict[str, Any]]] = {}


def reset_cache() -> None:
    _LOAD_CACHE.clear()


def load(path: Path | None = None) -> list[dict[str, Any]]:
    """Load the formula-trace records (memoized). ``[]`` when absent/unreadable/schema-mismatch (回帰ゼロ)."""
    out = Path(path) if path is not None else default_out_path()
    key = str(out)
    if key in _LOAD_CACHE:
        return _LOAD_CACHE[key]
    rows: list[dict[str, Any]] = []
    try:
        with open(out, encoding="utf-8") as handle:
            header = handle.readline()
            meta = json.loads(header) if header.strip() else {}
            if isinstance(meta, dict) and meta.get("schema") == SCHEMA \
                    and meta.get("version") == SCHEMA_VERSION:
                for line in handle:
                    line = line.strip()
                    if line:
                        rows.append(json.loads(line))
    except Exception:  # noqa: BLE001
        rows = []
    _LOAD_CACHE[key] = rows
    return rows


def docs_for_project(project_hint: str, *, path: Path | None = None) -> list[dict[str, Any]]:
    """Formula-trace records whose project matches ``project_hint`` (corporate-form / spacing insensitive)."""
    hint = owner_key(project_hint)
    rows = load(path)
    if not hint:
        return list(rows)
    return [r for r in rows if hint and (hint in owner_key(r.get("project", ""))
                                         or owner_key(r.get("project", "")) in hint)]


if __name__ == "__main__":
    summary = build()
    print(f"[build] xlsx_formula_trace records={summary['records']} "
          f"highlight_formulas={summary['report']['highlight_formulas']} "
          f"regressions={summary['report']['regressions']} -> {default_out_path()}")
