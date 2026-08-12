"""SOT-2680 — スケジュール/ID/体制のクロス参照ストア（cycle6 K4）.

案件ごとの ``02.計画/スケジュール.xlsx`` と ``00.提案/提案書*.pptx`` を **質問非依存** に走査し、
cycle6 で「クロス参照クエリ形が無く compute 前に予算切れ」だった 4 型を決定論ストアへ焼き込む:

* **ID 種別×件数の集計**（idx92）— 案件の非マークダウン出典から発行済みのマイルストーンID/タスクID/
  アクションID を種別ごとに distinct 集計（合計 = 3 種別の distinct 数の和）。
* **マイルストーン→タスク の逆引き**（idx94 素地）— タスク行の「関連マイルストーン」列とマイルストーン
  シートの「関連タスク」列の双方から ``MS→[T...]`` を構築（範囲表記 ``T05〜T08`` を展開）。
* **チェックポイント→タスク の派生クロス参照**（idx96）— CP 行の「関連MS/タスク」を、直接のタスク参照
  ∪（参照 MS の関連タスク）へ解決して related_task_ids に事前確定。
* **氏名→役職ロスター（案件スコープ厳守）**（idx94/72）— 提案書の体制表（``役職 → 氏名 → 責務``）から
  案件別に name→role を抽出。**案件横断の役割流用は禁止**（roster は必ずその案件の提案書由来）。

規律: 純粋な決定論抽出（openpyxl + docx/pptx 抽出のみ、LLM/genai 非使用 ⇒ cost $0）。``RAG_SCHEDULE_STORE``
既定 OFF ⇒ serve レーン／ツールは None（byte-identical）。build 自体は常に実行可能（べき等）。
"""
from __future__ import annotations

import io
import json
import os
import re
import unicodedata
from pathlib import Path
from typing import Any, Sequence

import openpyxl

from config import settings

from src.rag import corpus
from src.rag.corpus import FileRef, nfc
from src.rag.extract import office, passwords

SCHEMA = "schedule_store"
SCHEMA_VERSION = 1
_ON = {"1", "true", "yes", "on"}

_LOAD_CACHE: dict[str, list[dict[str, Any]]] = {}

# ID 記法（区切り非依存キーは norm_id）。タスク=T, マイルストーン=MS, アクション=A(会議録), CP=チェックポイント。
_TASK_RE = re.compile(r"^T\d{1,3}$")
_MS_RE = re.compile(r"^MS\d{1,3}$")
_CP_RE = re.compile(r"^CP\d{1,3}$")
_ACTION_RE = re.compile(r"\bA\d{2,3}\b")
# 提案書体制表の氏名（姓 名）。1〜5 文字ずつ、間に空白。
_NAME_RE = re.compile(r"^[一-龥ぁ-んァ-ヶ]{1,4}[ 　][一-龥ぁ-んァ-ヶ]{1,5}$")

# 提案書の体制表で使われる役職語彙（長い順に照合して部分一致の取りこぼしを防ぐ）。
_ROLE_VOCAB = (
    "エグゼクティブスポンサー",
    "リードデータサイエンティスト",
    "プロジェクトマネージャー",
    "プロジェクトマネージャ",
    "データサイエンティスト",
    "データエンジニア",
    "ビジネスアナリスト",
    "QAレビューアー",
    "QAレビュー担当",
    "QAレビューア",
    "推進運営",
)


def enabled() -> bool:
    """serve レーン／ツールを有効にするか。既定 OFF（``RAG_SCHEDULE_STORE``）⇒ byte-identical。"""
    return os.getenv("RAG_SCHEDULE_STORE", "0").strip().lower() in _ON


def default_out_path() -> Path:
    return settings.ARTIFACTS_DIR / "schedule_store.jsonl"


def default_report_path() -> Path:
    return settings.ARTIFACTS_DIR / "schedule_store_build_report.json"


# --------------------------------------------------------------------------- helpers
def name_key(value: Any) -> str:
    """氏名比較キー（空白除去 + NFKC）。'松本 真央' → 'まつもと' ではなく '松本真央'。"""
    return unicodedata.normalize("NFKC", str(value or "")).replace(" ", "").replace("　", "")


def _split_owners(value: Any) -> list[str]:
    if value is None:
        return []
    return [n.strip() for n in re.split(r"[、,/／;；]", str(value)) if n.strip()]


def expand_refs(value: Any, prefixes: Sequence[str] = ("MS", "CP", "T", "A")) -> list[str]:
    """'T05〜T08' / 'T05, T06' / 'MS2 / T12' 等を ID 列へ展開（範囲は端点含む inclusive）。"""
    if value is None:
        return []
    txt = nfc(str(value))
    out: list[str] = []
    # 範囲表記（Txx〜Tyy / MSx〜MSy）。桁数は左端点に合わせる。
    for pre in prefixes:
        for m in re.finditer(rf"{pre}(\d+)\s*[~〜～\-ー－]\s*{pre}?(\d+)", txt):
            a, b = int(m.group(1)), int(m.group(2))
            width = len(m.group(1))
            if 0 <= b - a <= 60:
                for k in range(a, b + 1):
                    out.append(f"{pre}{k:0{width}d}")
    # 単独 ID（MS/CP を T/A より先に照合し 'MS2' が 'S2' 等に割れないよう longest-first）。
    for m in re.finditer(r"(MS|CP|T|A)(\d+)", txt):
        out.append(f"{m.group(1)}{m.group(2)}")
    seen: set[str] = set()
    res: list[str] = []
    for x in out:
        if x not in seen:
            seen.add(x)
            res.append(x)
    return res


def _load_workbook(ref: FileRef):
    """暗号化 xlsx（かえで等）は passwords ヘルパで復号してから開く。"""
    if passwords.is_encrypted(ref.path):
        data = passwords.resolve(ref)
        if not data:
            raise ValueError(f"decrypt failed: {ref.rel}")
        return openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    return openpyxl.load_workbook(str(ref.path), data_only=True, read_only=True)


def _schedule_ref(project: str, refs: Sequence[FileRef]) -> FileRef | None:
    for r in refs:
        if r.project == project and r.category == "plan" and r.ext == "xlsx" \
                and nfc(r.name).startswith("スケジュール") and not nfc(r.name).startswith("~$"):
            return r
    return None


def _proposal_refs(project: str, refs: Sequence[FileRef]) -> list[FileRef]:
    return [r for r in refs if r.project == project and r.category == "proposal" and r.ext == "pptx"]


# --------------------------------------------------------------------------- roster (体制)
def _roster_from_proposals(project: str, refs: Sequence[FileRef]) -> list[dict[str, str]]:
    """提案書の体制表から case-scoped な name→role を抽出（役職トークンの直後の氏名トークン）。"""
    roster: dict[str, dict[str, str]] = {}
    for ref in _proposal_refs(project, refs):
        try:
            txt = office.extract_pptx(ref, None)
        except Exception:  # noqa: BLE001 — 抽出失敗は無視（ロスターは best-effort）
            continue
        tokens = [t.strip() for t in re.split(r"[\n\r\t]| {2,}", txt) if t.strip()]
        for i, tok in enumerate(tokens):
            role = _match_role(nfc(tok))
            if not role:
                continue
            for j in range(i + 1, min(i + 3, len(tokens))):
                cand = nfc(tokens[j])
                if _NAME_RE.match(cand):
                    key = name_key(cand)
                    roster.setdefault(key, {"role": role, "name": cand,
                                            "name_key": key, "source": ref.rel})
                    break
    return list(roster.values())


def _match_role(token: str) -> str | None:
    for role in _ROLE_VOCAB:
        if token == role or token.rstrip("ー") == role.rstrip("ー"):
            return role
    return None


# --------------------------------------------------------------------------- xlsx parse
def _cells(row) -> list[str]:
    return [nfc(str(c)) if c is not None else "" for c in row]


def _col(cells: list[str], *keywords: str) -> int | None:
    for j, c in enumerate(cells):
        if any(k in c for k in keywords):
            return j
    return None


def _parse_schedule(ref: FileRef) -> dict[str, Any]:
    """1 案件のスケジュール xlsx を tasks / milestones / checkpoints / ms_tasks へ正規化。"""
    tasks: dict[str, dict[str, Any]] = {}
    milestones: dict[str, dict[str, Any]] = {}
    checkpoints: dict[str, dict[str, Any]] = {}
    ms_tasks: dict[str, list[str]] = {}

    wb = _load_workbook(ref)
    for ws in wb.worksheets:
        rows = [list(r) for r in ws.iter_rows(values_only=True)]

        # (1) task sheet: header contains タスクID かつ直後に Txx が並ぶ。
        for hi, row in enumerate(rows):
            cells = _cells(row)
            if "タスクID" not in "|".join(cells):
                continue
            below = [(_cells(r2)) for r2 in rows[hi + 1: hi + 4]]
            if not any(_TASK_RE.match(c.strip()) for r2 in below for c in r2):
                continue
            ci_id = _col(cells, "タスクID")
            ci_name = _col(cells, "タスク名")
            ci_owner = _col(cells, "担当")
            ci_ms = _col(cells, "マイルストーン")
            ci_status = _col(cells, "ステータス", "状態")
            for r2 in rows[hi + 1:]:
                c2 = _cells(r2)
                if ci_id is None or ci_id >= len(c2):
                    continue
                tid = c2[ci_id].strip()
                if not _TASK_RE.match(tid):
                    continue
                owners = _split_owners(c2[ci_owner]) if ci_owner is not None and ci_owner < len(c2) else []
                tasks.setdefault(tid, {
                    "id": tid,
                    "name": c2[ci_name] if ci_name is not None and ci_name < len(c2) else "",
                    "owners": owners,
                    "owner_keys": [name_key(o) for o in owners],
                    "status": c2[ci_status] if ci_status is not None and ci_status < len(c2) else "",
                })
                if ci_ms is not None and ci_ms < len(c2):
                    for msid in (x for x in expand_refs(c2[ci_ms]) if x.startswith("MS")):
                        ms_tasks.setdefault(msid, [])
                        if tid not in ms_tasks[msid]:
                            ms_tasks[msid].append(tid)
            break

        # (2) milestone sheet: header has マイルストーンID / 'MS ID' + 関連タスク。
        for hi, row in enumerate(rows):
            cells = _cells(row)
            joined = "|".join(cells)
            if "マイルストーンID" not in joined and "MS ID" not in joined and "MSID" not in joined:
                continue
            ci_id = _col(cells, "マイルストーンID", "MS ID", "MSID")
            ci_name = _col(cells, "マイルストーン名", "概要", "名称")
            ci_rel = _col(cells, "関連タスク")
            for r2 in rows[hi + 1:]:
                c2 = _cells(r2)
                if ci_id is None or ci_id >= len(c2):
                    continue
                mid = c2[ci_id].strip()
                if not _MS_RE.match(mid):
                    continue
                rel = c2[ci_rel] if ci_rel is not None and ci_rel < len(c2) else ""
                rel_tasks = [x for x in expand_refs(rel) if x.startswith("T")]
                milestones.setdefault(mid, {
                    "id": mid,
                    "name": c2[ci_name] if ci_name is not None and ci_name < len(c2) else "",
                    "rel_tasks": rel_tasks,
                })
                for tid in rel_tasks:
                    ms_tasks.setdefault(mid, [])
                    if tid not in ms_tasks[mid]:
                        ms_tasks[mid].append(tid)
            break

        # (3) checkpoint rows: 「関連MS/タスク」列を持つ表で CPn 行を拾う。
        for hi, row in enumerate(rows):
            cells = _cells(row)
            joined = "|".join(cells)
            if "チェックポイント" not in joined and "関連MS" not in joined \
                    and not ("種別" in joined and "関連" in joined):
                continue
            ci_rel = _col(cells, "関連MS", "関連タスク", "関連")
            ci_content = _col(cells, "内容", "確認内容", "概要")
            for r2 in rows[hi + 1:]:
                c2 = _cells(r2)
                for cell in c2:
                    if _CP_RE.match(cell.strip()):
                        cpid = cell.strip()
                        rel = c2[ci_rel] if ci_rel is not None and ci_rel < len(c2) else ""
                        content = c2[ci_content] if ci_content is not None and ci_content < len(c2) else ""
                        checkpoints.setdefault(cpid, {"id": cpid, "content": content, "rel": rel})
                        break
    return {"tasks": tasks, "milestones": milestones, "checkpoints": checkpoints, "ms_tasks": ms_tasks}


def _action_ids(project: str, refs: Sequence[FileRef]) -> list[str]:
    """案件の非マークダウン出典（会議録/報告資料 docx・pptx）から発行済みアクションIDを distinct 抽出。"""
    ids: set[str] = set()
    for ref in refs:
        if ref.project != project or ref.category not in ("meeting", "report"):
            continue
        if ref.ext not in ("docx", "pptx"):
            continue
        try:
            if ref.ext == "docx":
                txt = office.raw_docx_text(ref.path)
            else:
                txt = office.extract_pptx(ref, None)
        except Exception:  # noqa: BLE001
            continue
        for m in _ACTION_RE.finditer(txt):
            ids.add(m.group(0))
    return sorted(ids, key=lambda s: (len(s), s))


def _resolve_checkpoint_tasks(cp: dict[str, Any], milestones: dict[str, Any]) -> list[str]:
    """CP の「関連MS/タスク」を tasks へ解決: 直接の T 参照 ∪（参照 MS の関連タスク）。"""
    refs = expand_refs(cp.get("rel"))
    tasks = [x for x in refs if x.startswith("T")]
    for msid in (x for x in refs if x.startswith("MS")):
        tasks.extend(milestones.get(msid, {}).get("rel_tasks", []))
    seen: set[str] = set()
    out: list[str] = []
    for t in tasks:
        if t not in seen:
            seen.add(t)
            out.append(t)
    return out


# --------------------------------------------------------------------------- build
def build_case(project: str, refs: Sequence[FileRef]) -> dict[str, Any] | None:
    sched = _schedule_ref(project, refs)
    if sched is None:
        return None
    parsed = _parse_schedule(sched)
    tasks = parsed["tasks"]
    milestones = parsed["milestones"]
    checkpoints = parsed["checkpoints"]
    ms_tasks = parsed["ms_tasks"]
    action_ids = _action_ids(project, refs)
    roster = _roster_from_proposals(project, refs)

    # チェックポイント→タスク を事前確定（質問非依存の派生クロス参照）。
    for cp in checkpoints.values():
        cp["related_task_ids"] = _resolve_checkpoint_tasks(cp, milestones)

    id_counts = {
        "task": len(tasks),
        "milestone": len(milestones) if milestones else _distinct_ms_from_tasks(ms_tasks, tasks),
        "action": len(action_ids),
    }
    # milestone シートが無い案件（単一シート型）は task 行の関連MS で distinct を数える。
    if not milestones:
        id_counts["milestone"] = len({m for m in ms_tasks})
    id_counts["total"] = id_counts["task"] + id_counts["milestone"] + id_counts["action"]

    return {
        "project": project,
        "schedule_file": sched.rel,
        "tasks": list(tasks.values()),
        "milestones": list(milestones.values()),
        "ms_tasks": ms_tasks,
        "checkpoints": list(checkpoints.values()),
        "roster": roster,
        "action_ids": action_ids,
        "id_counts": id_counts,
        "sources": {
            "schedule": sched.rel,
            "proposals": [r.rel for r in _proposal_refs(project, refs)],
        },
    }


def _distinct_ms_from_tasks(ms_tasks: dict[str, list[str]], tasks: dict[str, Any]) -> int:
    return len({m for m in ms_tasks})


def build(refs: Sequence[FileRef] | None = None, *, out: Path | None = None,
          write_report: bool = True) -> dict[str, Any]:
    """全案件のスケジュールクロス参照ストアを構築し JSONL へ書き出す（べき等・LLM 非使用）。"""
    all_refs = list(refs) if refs is not None else corpus.walk()
    projects = sorted({r.project for r in all_refs
                       if r.project and r.project not in ("社内管理", "案件横断", "")})
    records: list[dict[str, Any]] = []
    skipped: list[str] = []
    for project in projects:
        try:
            rec = build_case(project, all_refs)
        except Exception as exc:  # noqa: BLE001 — 1 案件の失敗が全体を壊さない
            skipped.append(f"{project}: {type(exc).__name__}: {exc}")
            continue
        if rec is None:
            skipped.append(f"{project}: no schedule")
            continue
        records.append(rec)

    out_path = out or default_out_path()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as fh:
        fh.write(json.dumps({"schema": SCHEMA, "version": SCHEMA_VERSION}, ensure_ascii=False) + "\n")
        for rec in records:
            fh.write(json.dumps(rec, ensure_ascii=False) + "\n")
    _LOAD_CACHE.pop(str(out_path), None)

    report = {
        "cases": len(records),
        "skipped": skipped,
        "totals": {rec["project"]: rec["id_counts"] for rec in records},
        "roster_sizes": {rec["project"]: len(rec["roster"]) for rec in records},
    }
    if write_report:
        with open(default_report_path(), "w", encoding="utf-8") as fh:
            json.dump(report, fh, ensure_ascii=False, indent=2)
    return {"records": len(records), "report": report}


def load(path: Path | None = None) -> list[dict[str, Any]]:
    """スケジュールストアを読み込む（memoized）。欠損/スキーマ不一致 ⇒ ``[]``（回帰ゼロ）。"""
    out = path or default_out_path()
    key = str(out)
    if key in _LOAD_CACHE:
        return _LOAD_CACHE[key]
    rows: list[dict[str, Any]] = []
    try:
        with open(out, encoding="utf-8") as fh:
            header = fh.readline()
            meta = json.loads(header) if header.strip() else {}
            if isinstance(meta, dict) and meta.get("schema") == SCHEMA \
                    and meta.get("version") == SCHEMA_VERSION:
                for line in fh:
                    line = line.strip()
                    if line:
                        rows.append(json.loads(line))
    except Exception:
        rows = []
    _LOAD_CACHE[key] = rows
    return rows


def case_record(project: str, *, path: Path | None = None) -> dict[str, Any] | None:
    """canonical 会社名に一致する 1 レコードを返す（無ければ None）。"""
    if not project:
        return None
    for rec in load(path):
        if rec.get("project") == project:
            return rec
    return None


if __name__ == "__main__":
    summary = build()
    print(json.dumps(summary["report"], ensure_ascii=False, indent=2))
