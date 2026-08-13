"""xlsx 可視事実（チャートメタデータ＋ハイライトセル全数）の build-time ストア（SOT-2653 / 事前計算事実層 追補 — クラスタB）.

Sonnet gold100 cycle4 の abstain のうち idx39（chart_read）/ idx65・idx97（highlight 由来の numeric）は、
証拠が「セルのテキスト」ではなく **xlsx の可視レイヤ**（チャートオブジェクトの系列→カラム対応・静的 fill と
条件付き書式で塗られたセルの全数・その色族要約）にあり、file_grep 全走査では到達できないのが主因だった
(`docs/ai/sonnet_cycle_analysis/cycle4.md` クラスタB)。同因の wrong idx42/77（黄セルの行列ラベル文脈）/
idx82（淡いオレンジ行を「存在しない」と誤答）も二次対象。

本モジュールは build 時に一度だけ、xlsx/xlsm を **質問を見ずに全数** 走査して

* **チャート** — 各グラフ（旧 ``c:`` チャート／新 ``cx:`` chartEx とも）のタイトルと、値系列が参照する
  カラムのヘッダ名（``_xlchart.v1.N`` 定義名や ``Sheet!$M$2:$M$…`` を解決してヘッダセルを読む）。
* **静的ハイライト** — 塗りつぶし(solid fill)されたセルを全数、``{sheet, 座標, 行, 列, 値, 色族, 行ラベル,
  列ヘッダ}`` 付きで収蔵。淡い色地(FFF0E6 等 低彩度)も色族へ分類する（champion の分類器が None に落とす
  淡橙行=idx82 を回収するため。ただし本ストア内限定で、serve path の分類器は不変）。
* **条件付き書式(CF dxf)ハイライト** — ``ws.conditional_formatting`` の ``cellIs`` ルールの色・演算子・
  しきい値・人間可読条件・該当セル。
* **派生構造** — 全数収蔵から質問非依存に導ける (a) 行全体ハイライト / (b) 列全体ハイライト /
  (c) 相関係数シート判定（全数値が [-1,1]）/ (d) 色族ごとの ``{n, min, max, range, 値サンプル}`` 要約。

を ``artifacts/visual_store.jsonl`` に焼く。serve path は事前計算済みを読むだけで **genai 呼び出しゼロ**。
読み出し/配線は :mod:`src.rag.agent.visual_lane`。

Design invariants（sibling の action_row_store / report_attr_store と同一）
--------------------------------------------------------------------------
* **Opt-in at serve time.** :func:`enabled` (``RAG_VISUAL_STORE``) が runtime 参照のみを gate する。
  default OFF ⇒ champion serve path は byte-identical。
* **Build は LLM フリー・追加的.** openpyxl の書式/定義名/CF を読むだけ。読めない文書は 1 件スキップし継続。
* **Question-independent.** universe は 03.データ/02.計画 等の xlsx。質問も gold も参照せず全チャート・全塗りを網羅。
* **No hardcoding.** gold 値・idx 番号を一切持たない。全値は原ファイルから読み、毎セル出典(doc/sheet/cell)付き。
* **Fail-open.** artifact 欠落・解析不能はすべて空へフォールバック（回帰ゼロ）。
"""
from __future__ import annotations

import colorsys
import io
import json
import os
import re
import unicodedata
import zipfile
from pathlib import Path
from typing import Any, Mapping, Sequence

from config import settings
from src.rag.corpus import FileRef, nfc, walk
from src.rag.extract import office as _office

SCHEMA = "visual-store"
SCHEMA_VERSION = 1

_ON = {"1", "true", "yes", "on"}

_XLSX_EXTS = {"xlsx", "xlsm"}

# A cell is a "full-line" highlight when its colour spans at least this fraction of the sheet's used
# columns (rows) — distinguishes a deliberately highlighted row/column from incidental single cells.
_LINE_FRACTION = 0.6
_LINE_MIN = 3

# DrawingML / chartEx namespaces.
_NS_A = "{http://schemas.openxmlformats.org/drawingml/2006/main}"
_NS_C = "{http://schemas.openxmlformats.org/drawingml/2006/chart}"
_NS_CX = "{http://schemas.microsoft.com/office/drawing/2014/chartex}"

_RANGE_RE = re.compile(r"^(?:'?([^'!]+)'?!)?\$?([A-Za-z]+)\$?(\d+)(?::\$?([A-Za-z]+)\$?(\d+))?$")
_GRAPH_NO = re.compile(r"(?:グラフ|チャート|graph|chart)\s*([0-9０-９]+)", re.IGNORECASE)


def enabled() -> bool:
    """True when the serve path may consult the visual store (default OFF — opt-in)."""
    return os.getenv("RAG_VISUAL_STORE", "0").strip().lower() in _ON


def pivot_context_enabled() -> bool:
    """True when highlight cells are enriched with pivot context at build time and the serve pivot
    lane may answer 抽出条件/集計内容 questions (SOT-2704, ``RAG_HIGHLIGHT_PIVOT_CONTEXT``; default OFF).

    OFF ⇒ ``_extract_sheet`` bakes no ``pivot_context`` key (store byte-identical) and the serve pivot
    lane defers, so both build artifact and answer path stay byte-identical to pre-SOT-2704.
    """
    return os.getenv("RAG_HIGHLIGHT_PIVOT_CONTEXT", "0").strip().lower() in _ON


def default_out_path() -> Path:
    return settings.ARTIFACTS_DIR / "visual_store.jsonl"


def default_report_path() -> Path:
    return settings.ARTIFACTS_DIR / "visual_store_build_report.json"


# --------------------------------------------------------------------------- colour classification
def classify_fill(fg: Any) -> str | None:
    """Coarse colour name for a solid fill's foreground colour, recovering **pale tints**.

    First defers to the shared, deterministic classifier (:func:`office._excel_color_name`) so a normal
    fill is named identically to the champion path. Only when that returns ``None`` does it retry a
    low-saturation tint (e.g. seashell ``FFF0E6`` → オレンジ): a pale but hue-bearing wash a human still
    reads as "オレンジ色にハイライト" (idx82). Pure greys / white / near-black banding stay ``None``.
    """
    name = _office._excel_color_name(fg)
    if name:
        return name
    rgb = _rgb_of(fg)
    if rgb is None:
        return _pale_tint(fg)
    return _pale_tint_from_rgb(rgb)


def _rgb_of(fg: Any) -> "tuple[int, int, int] | None":
    """Extract an ``(r, g, b)`` triple from an openpyxl Color (direct RGB only; indexed ⇒ None)."""
    argb = getattr(fg, "rgb", None)
    if not isinstance(argb, str):
        return None
    hexs = argb[-6:] if len(argb) >= 6 else ""
    if len(hexs) != 6:
        return None
    try:
        return int(hexs[0:2], 16), int(hexs[2:4], 16), int(hexs[4:6], 16)
    except ValueError:
        return None


def _pale_tint(fg: Any) -> str | None:
    rgb = _rgb_of(fg)
    return _pale_tint_from_rgb(rgb) if rgb is not None else None


def _pale_tint_from_rgb(rgb: "tuple[int, int, int]") -> str | None:
    """Name a light, low-saturation but hue-bearing wash; ``None`` for greys/white/near-black."""
    r, g, b = rgb
    h, s, v = colorsys.rgb_to_hsv(r / 255.0, g / 255.0, b / 255.0)
    # A tint the champion classifier drops: enough colour to read a hue (≥5%), still bright (≥85%).
    if s < 0.05 or v < 0.85:
        return None
    return _hue_name(h * 360.0)


def _hue_name(h: float) -> str:
    """Map a hue angle (deg) to the same coarse Japanese vocabulary the extractor emits."""
    if h < 15 or h >= 345:
        return "赤"
    if h < 40:
        return "オレンジ"
    if h < 70:
        return "黄"
    if h < 160:
        return "緑"
    if h < 200:
        return "水色"
    if h < 250:
        return "青"
    if h < 290:
        return "紫"
    return "ピンク"


# --------------------------------------------------------------------------- highlight extraction
def _col_letter(idx: int) -> str:
    from openpyxl.utils import get_column_letter
    return get_column_letter(idx)


def _cell_str(value: Any) -> str:
    return "" if value is None else str(value)


def _row_label(ws, row: int) -> Any:
    """The nearest non-empty label to the left on a highlighted cell's row (leftmost id column)."""
    for c in range(1, min(ws.max_column, 8) + 1):
        v = ws.cell(row, c).value
        if v is not None and str(v).strip() != "":
            return v
    return None


# --------------------------------------------------------------------------- pivot context (SOT-2704)
# Aggregate/value-column headers in a spreadsheet pivot: a group-by (dimension) column never carries
# these markers, so a left column whose header matches is a value column, not an extraction condition.
_AGG_MARK = re.compile(
    r"平均|合計|総計|小計|個数|件数|最大|最小|中央|分散|標準偏差|割合|比率|構成比|"
    r"count|sum|avg|average|mean|median|max|min|std|total|/", re.IGNORECASE)
# The leading aggregation token of a value-column header (``平均 / bmi`` → ``平均``).
_AGG_LEAD = re.compile(
    r"^\s*(平均|合計|総計|小計|個数|件数|最大|最小|中央値?|分散|標準偏差|割合|比率|"
    r"average|sum|count|mean|median|max|min|std|total)\s*/", re.IGNORECASE)


def _fmt_label(v: Any) -> str:
    """A pivot label as text, normalising integral floats (``2.0`` → ``2``) so a numeric group key
    like ``charges=2`` renders exactly as the sheet displays it."""
    if isinstance(v, bool):
        return str(v)
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    if isinstance(v, int):
        return str(v)
    return str(v).strip()


def _detect_header_row(ws, limit: int = 25) -> "int | None":
    """The pivot's header row: the first row (top-down) carrying ≥2 non-empty *string* labels. In a
    pandas-exported pivot the header sits above the data (rows above it are blank); a plain table has
    it at row 1. Returns None when no label row is found (⇒ the sheet is not treated as a pivot)."""
    for r in range(1, min(ws.max_row or 0, limit) + 1):
        vals = [ws.cell(r, c).value for c in range(1, (ws.max_column or 0) + 1)]
        strs = [v for v in vals if isinstance(v, str) and v.strip() != ""]
        if len(strs) >= 2:
            return r
    return None


def _pivot_context(ws, row: int, col: int, header_row: "int | None") -> "dict[str, Any] | None":
    """Question-independent extraction condition + aggregation for a highlighted pivot cell.

    Reads the cell's own column header (``column_header`` — the 集計内容, e.g. ``平均 / bmi``) and, for
    each *dimension* column to its left (a left column whose header is NOT an aggregate marker), the
    nearest non-empty label at-or-above ``row`` (upward scan) — the parent group key of that cell. For
    ``F22`` in a sex/smoker/region/charges pivot this yields ``sex=female, smoker=yes, region=southeast,
    charges=2``. Returns None (⇒ no enrichment) when the sheet is not a pivot below the header or no
    dimension binds. Never reads the question or gold.
    """
    if header_row is None or header_row >= row:
        return None
    raw_header = ws.cell(header_row, col).value
    if raw_header is None or str(raw_header).strip() == "":
        return None
    column_header = str(raw_header).strip()
    row_context: list[dict[str, Any]] = []
    for c in range(1, col):
        h = ws.cell(header_row, c).value
        if h is None or str(h).strip() == "":
            continue
        name = str(h).strip()
        if _AGG_MARK.search(name):  # a value/aggregate column, not an extraction condition
            continue
        label = None
        for rr in range(row, header_row, -1):  # upward scan for the nearest parent label
            v = ws.cell(rr, c).value
            if v is not None and str(v).strip() != "":
                label = v
                break
        if label is None:
            continue
        row_context.append({"name": name, "value": _fmt_label(label)})
    if not row_context:
        return None
    m = _AGG_LEAD.match(column_header)
    return {
        "header_row": header_row,
        "column_header": column_header,
        "agg_hint": m.group(1) if m else None,
        "row_context": row_context,
    }


def _extract_sheet(ws) -> dict[str, Any]:
    """Structured visual facts for one worksheet (static fills + derived line/colour structures)."""
    used_cols = ws.max_column or 0
    used_rows = ws.max_row or 0
    statics: list[dict[str, Any]] = []
    numeric_vals: list[float] = []
    by_row: dict[int, dict[str, int]] = {}
    by_col: dict[int, dict[str, int]] = {}
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                numeric_vals.append(float(v))
            fill = cell.fill
            if not (fill and fill.patternType):
                continue
            color = classify_fill(getattr(fill, "fgColor", None))
            if not color:
                continue
            header = ws.cell(1, cell.column).value
            statics.append({
                "cell": cell.coordinate, "row": cell.row, "column": cell.column,
                "col_letter": _col_letter(cell.column), "value": _cell_str(v), "color": color,
                "col_header": None if header in (None, "") else str(header),
            })
            by_row.setdefault(cell.row, {})[color] = by_row.setdefault(cell.row, {}).get(color, 0) + 1
            by_col.setdefault(cell.column, {})[color] = by_col.setdefault(cell.column, {}).get(color, 0) + 1

    row_hi = _line_highlights(by_row, used_cols, axis="row")
    col_hi = _line_highlights(by_col, used_rows, axis="col")
    # Attach per-highlighted-row the cell values keyed by header, so a "行のタスクIDをすべて" question can
    # read the row's ID column deterministically (idx82).
    for rh in row_hi:
        rh["cells"] = [{"col_letter": _col_letter(c), "header": _hdr(ws, c), "value": _cell_str(ws.cell(rh["row"], c).value)}
                       for c in range(1, used_cols + 1)]
    for ch in col_hi:
        ch["col_letter"] = _col_letter(ch["col"])
        ch["header"] = _hdr(ws, ch["col"])

    # SOT-2704: bake each highlighted cell's pivot context (parent extraction conditions + the
    # aggregation column header) so serve can answer 抽出条件/集計内容 without in-budget reconstruction.
    # Gated so OFF ⇒ no new key ⇒ store + serve byte-identical.
    if pivot_context_enabled() and statics:
        header_row = _detect_header_row(ws)
        for it in statics:
            pc = _pivot_context(ws, it["row"], it["column"], header_row)
            if pc:
                it["pivot_context"] = pc

    return {
        "state": ws.sheet_state,
        "used_rows": used_rows, "used_cols": used_cols,
        "is_correlation": _is_correlation(numeric_vals),
        "static": statics,
        "cf": _cf_items(ws),
        "row_highlights": row_hi,
        "col_highlights": col_hi,
        "color_summary": _color_summary(statics),
    }


def _hdr(ws, col: int) -> Any:
    v = ws.cell(1, col).value
    return None if v in (None, "") else str(v)


def _line_highlights(by_line: Mapping[int, Mapping[str, int]], span: int, *, axis: str) -> list[dict[str, Any]]:
    """Rows/cols whose SAME colour spans a majority of the used span — a deliberate line highlight."""
    thr = max(_LINE_MIN, int(span * _LINE_FRACTION)) if span else _LINE_MIN
    out: list[dict[str, Any]] = []
    key = "row" if axis == "row" else "col"
    for line, colors in sorted(by_line.items()):
        for color, n in colors.items():
            if n >= thr:
                out.append({key: line, "color": color, "n": n})
    return out


def _is_correlation(vals: Sequence[float]) -> bool:
    """A correlation-matrix sheet: many numeric cells, all within [-1, 1] (incl. the 1.0 diagonal)."""
    if len(vals) < 8:
        return False
    return all(-1.0001 <= v <= 1.0001 for v in vals)


def _color_summary(statics: Sequence[Mapping[str, Any]]) -> dict[str, Any]:
    """Per colour family: count + numeric min/max/range over the highlighted cells (idx65/97 material)."""
    out: dict[str, Any] = {}
    for it in statics:
        color = it["color"]
        num = _to_num(it["value"])
        slot = out.setdefault(color, {"n": 0, "numeric": []})
        slot["n"] += 1
        if num is not None:
            slot["numeric"].append(num)
    for color, slot in out.items():
        nums = slot.pop("numeric")
        if nums:
            slot["min"] = min(nums)
            slot["max"] = max(nums)
            slot["range"] = round(abs(max(nums) - min(nums)), 6)
            slot["n_numeric"] = len(nums)
    return out


def _to_num(text: Any) -> "float | None":
    try:
        return float(str(text).replace(",", "").strip())
    except (TypeError, ValueError):
        return None


# --------------------------------------------------------------------------- conditional formatting
_CF_OPS = {"lessThan": "<", "lessThanOrEqual": "<=", "greaterThan": ">", "greaterThanOrEqual": ">=",
           "equal": "=", "notEqual": "!=", "between": "の範囲", "notBetween": "の範囲外",
           "containsText": "を含む"}


def _cf_condition_text(operator: str | None, formulas: Any) -> str:
    vals = ", ".join(str(x) for x in (formulas or []))
    label = _CF_OPS.get(operator or "", operator or "")
    if operator in ("between", "notBetween"):
        return f"セルの値が {vals} {label}".strip()
    return f"セルの値 {label} {vals}".strip()


def _cf_items(ws) -> list[dict[str, Any]]:
    """``cellIs`` conditional-format rules: colour (dxf fill), operator, threshold, human-readable cond."""
    out: list[dict[str, Any]] = []
    try:
        cf = ws.conditional_formatting
    except Exception:  # noqa: BLE001
        return out
    for rng in cf:
        for rule in cf[rng]:
            if getattr(rule, "type", None) != "cellIs":
                continue
            dxf = getattr(rule, "dxf", None)
            fill = getattr(dxf, "fill", None) if dxf else None
            color = None
            if fill is not None:
                for attr in ("bgColor", "fgColor"):
                    c = getattr(fill, attr, None)
                    nm = classify_fill(c) if c is not None else None
                    if nm:
                        color = nm
                        break
            formulas = [str(f) for f in (getattr(rule, "formula", None) or [])]
            out.append({
                "color": color, "operator": rule.operator, "formula": formulas,
                "condition": _cf_condition_text(rule.operator, rule.formula), "sqref": str(rng.sqref),
            })
    return out


# --------------------------------------------------------------------------- chart extraction
def _rich_text(el) -> str:
    """Concatenate the run texts (``<a:t>``) inside a chart title element, NFC-normalized."""
    parts = [t.text or "" for t in el.iter(_NS_A + "t")]
    return nfc("".join(parts)).strip()


def _resolve_defined_names(zf: zipfile.ZipFile) -> dict[str, str]:
    """``defined name → range string`` from xl/workbook.xml (e.g. ``_xlchart.v1.3`` → ``train!$M$2:$M$8646``)."""
    out: dict[str, str] = {}
    try:
        wb = zf.read("xl/workbook.xml").decode("utf-8", "replace")
    except KeyError:
        return out
    for m in re.finditer(r"<definedName[^>]*\bname=\"([^\"]+)\"[^>]*>(.*?)</definedName>", wb, re.S):
        out[m.group(1)] = m.group(2).strip()
    return out


def _range_column(range_str: str) -> "tuple[str | None, str, int] | None":
    """Parse ``Sheet!$M$2:$M$…`` → ``(sheet, 'M', start_row)`` (first column & row only)."""
    m = _RANGE_RE.match(range_str.strip())
    if not m:
        return None
    return m.group(1), m.group(2).upper(), int(m.group(3))


def _column_header(wb, sheet: str | None, col_letter: str, start_row: int) -> "tuple[str | None, str | None]":
    """The header of a data column: the cell one row above the value range's first row. Returns (header, sheet)."""
    from openpyxl.utils import column_index_from_string
    ws = None
    if sheet and sheet in wb.sheetnames:
        ws = wb[sheet]
    elif sheet:
        for name in wb.sheetnames:
            if nfc(name) == nfc(sheet):
                ws = wb[name]
                break
    if ws is None:
        return None, sheet
    header_row = start_row - 1 if start_row > 1 else 1
    try:
        val = ws.cell(header_row, column_index_from_string(col_letter)).value
    except Exception:  # noqa: BLE001
        return None, ws.title
    return (None if val in (None, "") else str(val)), ws.title


def _charts(data: bytes, wb) -> list[dict[str, Any]]:
    """Every chart's title + the visualised value column header (chartEx defined-name / classic range)."""
    import xml.etree.ElementTree as ET
    charts: list[dict[str, Any]] = []
    try:
        zf = zipfile.ZipFile(io.BytesIO(data))
    except Exception:  # noqa: BLE001
        return charts
    with zf:
        defnames = _resolve_defined_names(zf)
        members = sorted(n for n in zf.namelist()
                         if re.search(r"charts/(chart|chartEx)\d+\.xml$", n))
        for member in members:
            try:
                root = ET.fromstring(zf.read(member))
            except Exception:  # noqa: BLE001
                continue
            title, refs = _chart_title_and_val_refs(root)
            columns: list[dict[str, Any]] = []
            for ref in refs:
                rng = defnames.get(ref, ref)  # chartEx `_xlchart.v1.N` → range; classic ref already a range
                parsed = _range_column(rng)
                if not parsed:
                    continue
                sheet, col, start_row = parsed
                header, resolved_sheet = _column_header(wb, sheet, col, start_row)
                columns.append({"ref": rng, "column": col, "header": header, "data_sheet": resolved_sheet})
            if title or columns:
                charts.append({"member": member.rsplit("/", 1)[-1], "title": title, "columns": columns})
    return charts


def _chart_title_and_val_refs(root) -> "tuple[str | None, list[str]]":
    """Title text + the value-dimension references for a chart (both chartEx ``cx`` and classic ``c``)."""
    tag = root.tag
    title = None
    refs: list[str] = []
    if tag.endswith("chartSpace") and _NS_CX in tag:
        for t in root.iter(_NS_CX + "title"):
            title = _rich_text(t) or title
            break
        for data in root.iter(_NS_CX + "data"):
            for dim in data:
                if dim.tag == _NS_CX + "numDim" and dim.get("type") == "val":
                    f = dim.find(_NS_CX + "f")
                    if f is not None and f.text:
                        refs.append(f.text.strip())
    else:
        for t in root.iter(_NS_C + "title"):
            title = _rich_text(t) or title
            break
        for ser in root.iter(_NS_C + "ser"):
            val = ser.find(_NS_C + "val")
            if val is not None:
                for f in val.iter(_NS_C + "f"):
                    if f.text:
                        refs.append(f.text.strip())
                        break
    return title, refs


# --------------------------------------------------------------------------- per-document record
def compute_doc(ref: FileRef) -> dict[str, Any] | None:
    """Build one document's visual record (``None`` when nothing visual is present — no fabrication)."""
    import openpyxl
    data = ref.path.read_bytes()
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    sheets: dict[str, Any] = {}
    for ws in wb.worksheets:
        s = _extract_sheet(ws)
        if s["static"] or s["cf"] or s["row_highlights"] or s["col_highlights"]:
            sheets[ws.title] = s
    charts = _charts(data, wb)
    if not sheets and not charts:
        return None
    return {
        "doc_id": nfc(ref.rel), "project": nfc(ref.project), "doc_name": nfc(ref.name),
        "sheets": sheets, "charts": charts,
    }


# --------------------------------------------------------------------------- universe / build / io
def _universe(refs: Sequence[FileRef]) -> list[FileRef]:
    out: list[FileRef] = []
    for r in refs:
        if r.ext not in _XLSX_EXTS:
            continue
        name = nfc(r.name).lower()
        if name.startswith("~$") or "old" in name:
            continue
        out.append(r)
    out.sort(key=lambda r: r.rel)
    return out


def write_store(records: Sequence[Mapping[str, Any]], path: Path | None = None) -> dict[str, int]:
    """Atomically write a reproducible JSONL store (schema header + doc_id-sorted rows)."""
    out = Path(path) if path is not None else default_out_path()
    out.parent.mkdir(parents=True, exist_ok=True)
    ordered = sorted(records, key=lambda r: r.get("doc_id", ""))
    tmp = out.with_suffix(out.suffix + ".tmp")
    with tmp.open("w", encoding="utf-8", newline="\n") as handle:
        handle.write(json.dumps({"schema": SCHEMA, "version": SCHEMA_VERSION},
                                ensure_ascii=False, sort_keys=True) + "\n")
        for rec in ordered:
            handle.write(json.dumps(rec, ensure_ascii=False, sort_keys=True) + "\n")
    tmp.replace(out)
    reset_cache()
    return {"docs": len(ordered)}


def build(refs: Sequence[FileRef] | None = None, *, out: Path | None = None,
          write_report: bool = True) -> dict[str, Any]:
    """Scan every xlsx and bake the visual store (question-independent, LLM-free, fail-open)."""
    refs = list(refs) if refs is not None else list(walk())
    universe = _universe(refs)
    records: list[dict[str, Any]] = []
    skipped: list[str] = []
    for ref in universe:
        try:
            rec = compute_doc(ref)
        except Exception:  # noqa: BLE001 — one bad workbook must not sink the build
            rec = None
        if rec is not None:
            records.append(rec)
        else:
            skipped.append(nfc(ref.rel))
    stats = write_store(records, out)
    report = {
        "schema": SCHEMA, "version": SCHEMA_VERSION,
        "universe": len(universe), "records": len(records), "skipped": skipped,
        "charts": sum(len(r.get("charts", [])) for r in records),
        "highlighted_cells": sum(sum(len(s.get("static", [])) for s in r.get("sheets", {}).values())
                                 for r in records),
        "cf_rules": sum(sum(len(s.get("cf", [])) for s in r.get("sheets", {}).values())
                        for r in records),
    }
    if write_report:
        rp = default_report_path()
        rp.parent.mkdir(parents=True, exist_ok=True)
        rp.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"records": len(records), "docs": stats["docs"], "report": report}


# --------------------------------------------------------------------------- load / read (minimal API)
_LOAD_CACHE: dict[str, list[dict[str, Any]]] = {}


def reset_cache() -> None:
    _LOAD_CACHE.clear()


def load(path: Path | None = None) -> list[dict[str, Any]]:
    """Load the visual records (memoized). ``[]`` when absent/unreadable/schema-mismatch (回帰ゼロ)."""
    out = Path(path) if path is not None else default_out_path()
    key = str(out)
    if key in _LOAD_CACHE:
        return _LOAD_CACHE[key]
    rows: list[dict[str, Any]] = []
    try:
        with open(out, encoding="utf-8") as handle:
            header = handle.readline()
            meta = json.loads(header) if header.strip() else {}
            if isinstance(meta, dict) and meta.get("schema") == SCHEMA \
                    and meta.get("version") == SCHEMA_VERSION:
                for line in handle:
                    line = line.strip()
                    if line:
                        rows.append(json.loads(line))
    except Exception:  # noqa: BLE001
        rows = []
    _LOAD_CACHE[key] = rows
    return rows


# 案件名の識別キー（法人格・空白・ケースを吸収）— action_row_store と同じ規約。
_CORP = r"(?:株式会社|医療法人社団|一般社団法人|一般財団法人|有限会社|合同会社|合資会社)"
_CORP_PREFIX = re.compile(rf"^{_CORP}\s*")
_CORP_SUFFIX = re.compile(rf"\s*{_CORP}$")


def owner_key(value: Any) -> str:
    s = unicodedata.normalize("NFKC", str(value or ""))
    s = _CORP_PREFIX.sub("", s)
    s = _CORP_SUFFIX.sub("", s)
    return re.sub(r"[\s　]", "", s).lower()


def docs_for_project(project_hint: str, *, path: Path | None = None) -> list[dict[str, Any]]:
    """Visual records whose project matches ``project_hint`` (corporate-form / spacing insensitive)."""
    hint = owner_key(project_hint)
    rows = load(path)
    if not hint:
        return list(rows)
    return [r for r in rows if hint and (hint in owner_key(r.get("project", ""))
                                         or owner_key(r.get("project", "")) in hint)]


if __name__ == "__main__":
    summary = build()
    print(json.dumps(summary["report"], ensure_ascii=False, indent=2))
