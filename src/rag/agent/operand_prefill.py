"""SOT-2616 — NUMERIC operand candidate prefill (guess-and-check → 選択問題).

Phase-0 diagnostics (``docs/ai/budget32_trace_classification.md``) showed the NUMERIC route's
``BUDGET_EXHAUSTED`` losses are not a *calculation* deficit — the PoT hard lane (SOT-2586) can already
compute exactly — but an *operand delivery* deficit: ``derivation`` 88 試行 ok 36 (41%), with the LLM
推測→compute→失敗→再推測 (idx47 compute×7, idx57/95 compute×8), and idx76 spending 18 ターン中 17 回を
search したまま operand をひとつも束ねられない。The missing piece is **listing the operand candidates
deterministically before the loop**, so the LLM's job shrinks from 自由探索 to *選択* (どの値が
base/rate/条件か) — the same "事前確定情報を増やして到達率を上げる" axis Wave A used to recover 12 ABSTAIN→
MATCH, and explicitly *not* the firing-condition relaxation that crashed SOT-2601 to net12.

What this module does
---------------------
Given a question and the registry-resolved target documents, it enumerates **numeric cell candidates**
(value・unit・cell 座標・sheet・document) from the existing offline assets — the SOT-2533 structure store's
highlighted cells and the resolved spreadsheets' SOT-2531 normalized rows — ranks them by relevance to the
question's metric/entity tokens, caps the list (default 12), and hands them to the Evidence Packet as an
``operand_candidates`` catalog with stable ids. The LLM then *selects* from the catalog and the PoT lane
(:func:`src.rag.agent.pot_lane.resolve_operand_selections`) binds the chosen ids as sourced operands —
value/unit/source come from the catalog, never re-typed by the model.

Design invariants (mirroring the sibling opt-ins — evidence_packet / pot_lane / structure_store):
  * **Opt-in, byte-identical OFF.**  :func:`enabled` (``RAG_OPERAND_PREFILL``) gates the whole thing; the
    default-OFF serve path never enumerates, injects nothing, and is byte-identical.
  * **No degradation when empty.**  A question whose resolved documents yield *no* numeric candidate
    injects nothing and follows the ordinary path — the prefill only ever *adds* a catalog, never removes
    a lane or forces an approximation.
  * **Deterministic & answer-free.**  Every candidate is a real cell's value with its provenance; the
    module never invents a value and never chooses the answer. Selection + arithmetic stay downstream.
"""
from __future__ import annotations

import os
import re
from dataclasses import dataclass, field
from typing import Any, Iterable, Sequence

from src.rag.index import structure_store as _store

_ON = {"1", "true", "yes", "on"}

# The catalog never grows without bound: a NUMERIC derivation legitimately needs a handful of operands, so
# a dozen well-ranked candidates keep the selection tractable for the LLM without drowning it in the whole
# sheet. Ceiling is tunable via the env for focused experiments.
_DEFAULT_MAX = 12

# Row cap when scanning a resolved spreadsheet's normalized rows for candidates — the same conservative
# window the office normalizer uses, kept small so a wide sheet cannot dominate a single question's budget.
_ROW_SCAN_CAP = 200

_ASCII_TOK = re.compile(r"[A-Za-z0-9_]{2,}")
_JP_RUN = re.compile(r"[ぁ-んァ-ヶ一-鿿々ー]{2,}")


def enabled() -> bool:
    """True when the NUMERIC serve path should prefill operand candidates (default OFF — opt-in)."""
    return os.getenv("RAG_OPERAND_PREFILL", "0").strip().lower() in _ON


def max_candidates() -> int:
    """The candidate ceiling (``RAG_OPERAND_PREFILL_MAX`` override; default :data:`_DEFAULT_MAX`)."""
    raw = os.getenv("RAG_OPERAND_PREFILL_MAX", "").strip()
    if raw:
        try:
            n = int(raw)
            if n > 0:
                return n
        except ValueError:
            pass
    return _DEFAULT_MAX


# --------------------------------------------------------------------------- candidate model
@dataclass(frozen=True)
class OperandCandidate:
    """One deterministically-enumerated numeric operand candidate with full provenance.

    ``id`` is the stable handle the LLM selects by (``op1``…), so the chosen value/unit/source are copied
    from the catalog by the PoT handoff rather than re-typed. ``source`` is the ``doc:sheet!cell`` string
    the lane's operand layer requires; ``label`` is the human name (row/column/group context) that makes
    the candidate選べる.
    """

    id: str
    value: Any                 # the raw cell value, verbatim (str/number) — the lane re-parses exactly
    number: float              # its numeric reading, used only for ranking/dedup (never committed)
    label: str
    unit: str
    source: str
    sheet: str = ""
    cell: str = ""
    doc: str = ""
    origin: str = ""           # "highlight" | "row" — where the candidate was surfaced from
    context: dict[str, Any] = field(default_factory=dict)

    def to_catalog_entry(self) -> dict[str, Any]:
        """The JSON entry injected into the Evidence Packet / passed back to the PoT handoff."""
        return {
            "id": self.id,
            "value": self.value,
            "unit": self.unit,
            "source": self.source,
            "label": self.label,
            "sheet": self.sheet,
            "cell": self.cell,
            "doc": self.doc,
            "origin": self.origin,
        }


# --------------------------------------------------------------------------- tokenisation / ranking
def _tokens(text: str) -> list[str]:
    out: list[str] = []
    for m in list(_ASCII_TOK.finditer(text)) + list(_JP_RUN.finditer(text)):
        tok = m.group(0).lower()
        if tok and tok not in out:
            out.append(tok)
    return out


def _overlap(qtokens: Sequence[str], label: str, extra: Iterable[str]) -> int:
    """Token overlap between the question and a candidate's label + structural context (sheet/column)."""
    hay = set(_tokens(label))
    for e in extra:
        hay |= set(_tokens(str(e)))
    return sum(1 for t in qtokens if t in hay)


# --------------------------------------------------------------------------- enumeration sources
def _highlight_candidates(rel: str) -> list[dict[str, Any]]:
    """Numeric highlighted cells for a resolved doc (offline, from the structure store)."""
    try:
        return _store.stored_numeric_cells(rel)
    except Exception:  # noqa: BLE001 — additive; a store hiccup must never break the serve path
        return []


def _label_for_group(group: dict[str, Any] | None, sheet: str, cell: str) -> str:
    if group:
        return " / ".join(f"{k}={v}" for k, v in group.items())
    return f"{sheet}!{cell}" if sheet or cell else cell


def _row_candidates(rel: str) -> list[dict[str, Any]]:
    """Numeric cells from a resolved spreadsheet's normalized rows (best-effort, live).

    Reads the same conservative forward-filled rows the fact/aggregate pipelines use, naming each numeric
    cell by its row's leading text label + column header so a candidate is選べる. Wrapped whole in a guard:
    a missing corpus / encrypted file / openpyxl absence degrades to the highlight-only layer, never an
    error. Only runs when the prefill is enabled (callers gate), so the champion path never opens a file.
    """
    out: list[dict[str, Any]] = []
    try:
        from openpyxl.utils import get_column_letter

        from src.rag.extract.office import normalized_xlsx_rows
        from src.rag.tools.extract_tools import resolve_ref
        import openpyxl

        ref = resolve_ref(rel)
        if ref.ext not in {"xlsx", "xlsm"} or ref.name.startswith("~$"):
            return []
        wb = openpyxl.load_workbook(str(ref.path), data_only=True, read_only=True)
    except Exception:  # noqa: BLE001
        return []
    try:
        for ws in wb.worksheets:
            try:
                rows = normalized_xlsx_rows(ws, max_rows=_ROW_SCAN_CAP)
            except Exception:  # noqa: BLE001
                continue
            if not rows:
                continue
            header = _densest_row(rows)
            headers = _row_headers(header)
            ordinal_cols = {ci for ci, h in headers.items() if _is_ordinal_header(h)}
            for nrow in rows:
                if nrow is header:
                    continue
                row_label = _row_leading_label(nrow)
                for ci, value in enumerate(nrow.values):
                    num = _store._as_number(value)
                    if num is None:
                        continue
                    # Drop identifier/ordinal columns and pure row-index cells: a "No."/"id"/"番号"
                    # column value or a cell that equals its own row number is a record index, not an
                    # operand. This keeps the candidate menu to real data/metric cells.
                    if ci in ordinal_cols or (num == float(nrow.row) and not headers.get(ci)):
                        continue
                    col = headers.get(ci, "")
                    coord = f"{get_column_letter(ci + 1)}{nrow.row}"
                    name = " ".join(p for p in (row_label, col) if p) or coord
                    out.append({
                        "value": value, "number": num, "label": name,
                        "sheet": ws.title, "cell": coord, "column": col,
                        "doc": rel, "row_label": row_label,
                    })
    finally:
        try:
            wb.close()
        except Exception:  # noqa: BLE001
            pass
    return out


def _densest_row(rows):
    scan = rows[:30]
    return max(scan, key=lambda r: sum(1 for v in r.values if v is not None and str(v).strip()))


def _row_headers(header) -> dict[int, str]:
    return {ci: str(v).strip() for ci, v in enumerate(header.values)
            if v is not None and str(v).strip()}


# Column headers that mark a record identifier / ordinal, not an operand (folded, exact or suffix match).
_ORDINAL_HEADERS = frozenset({"no", "no.", "no．", "№", "id", "index", "num", "#",
                              "番号", "項番", "通し番号", "連番", "行番号"})


def _is_ordinal_header(header: str) -> bool:
    h = header.strip().lower().replace(" ", "")
    if not h:
        return False
    if h in _ORDINAL_HEADERS:
        return True
    # trailing "…No"/"…No."/"…id" (e.g. "タスクNo." / "マイルストーンNo.") — an ordinal column with a prefix.
    return h.endswith(("no.", "no．", "no")) and len(h) <= 12 or h.endswith("id") and len(h) <= 8


def _row_leading_label(nrow) -> str:
    for v in nrow.values:
        if v is not None and str(v).strip() and _store._as_number(v) is None:
            return str(v).strip()
    return ""


# --------------------------------------------------------------------------- project-scoped fallback
# When the registry resolves a project but no specific document (idx76: "AOMINE" 別名 — the numeric
# operands live in the project's data spreadsheet, but the question names it by an alias the resolver
# can't match), fall back to the project's own xlsx/xlsm files so the operands are still pre-listed.
_FALLBACK_FILE_CAP = 3


def _project_scoped_rels(project: str | None) -> list[str]:
    """Data spreadsheets (xlsx/xlsm) in ``project`` from the registry, data-ish files first (capped)."""
    if not project:
        return []
    try:
        from src.rag.index import document_registry as _dr
        rels = _dr.rels_for_project(project, extensions={"xlsx", "xlsm"})
    except Exception:  # noqa: BLE001
        return []

    def _rank(rel: str) -> tuple[int, str]:
        low = rel.lower()
        pref = 0 if ("データ" in rel or "train" in low or "data" in low) else 1
        return (pref, rel)

    return sorted(dict.fromkeys(rels), key=_rank)[:_FALLBACK_FILE_CAP]


# --------------------------------------------------------------------------- public: enumerate
def enumerate_candidates(question: str, resolved_docs: Sequence[str], *,
                         project: str | None = None,
                         limit: int | None = None) -> list[OperandCandidate]:
    """Rank and cap the numeric operand candidates for ``question`` over its resolved documents.

    ``resolved_docs`` = the registry-resolved NFC rel paths (from the Evidence Packet). Candidates are
    drawn from the structure store's highlighted cells (offline) and the resolved spreadsheets' normalized
    rows (best-effort live), de-duplicated by ``(doc,sheet,cell)``, ranked by question-token overlap (then
    a stable source order), and truncated to ``limit`` (default :func:`max_candidates`). When the resolved
    documents yield nothing numeric *and* a ``project`` is known, it falls back to that project's data
    spreadsheets (idx76-type alias questions). Returns ``[]`` when nothing numeric is reachable — the
    caller then injects no catalog (no degradation).
    """
    cap = limit if limit is not None else max_candidates()
    qtokens = _tokens(question)
    raw: list[dict[str, Any]] = []
    seen: set[tuple[str, str, str]] = set()

    def _push(rec: dict[str, Any], origin: str) -> None:
        sheet = str(rec.get("sheet", "") or "")
        cell = str(rec.get("cell", "") or "")
        doc = str(rec.get("doc", "") or "")
        key = (doc, sheet, cell)
        if cell and key in seen:
            return
        if cell:
            seen.add(key)
        rec = dict(rec)
        rec["origin"] = origin
        raw.append(rec)

    def _enumerate_rels(rels: Sequence[str]) -> None:
        for rel in dict.fromkeys(rels):
            for rec in _highlight_candidates(rel):
                rec.setdefault("label", _label_for_group(rec.get("group"), rec.get("sheet", ""),
                                                         rec.get("cell", "")))
                _push(rec, "highlight")
            for rec in _row_candidates(rel):
                _push(rec, "row")

    _enumerate_rels(resolved_docs)
    if not raw and project:
        _enumerate_rels(_project_scoped_rels(project))

    if not raw:
        return []

    def _rank_key(rec: dict[str, Any]):
        label = str(rec.get("label", ""))
        extra = (rec.get("sheet", ""), rec.get("column", "") or "", rec.get("doc", ""))
        ov = _overlap(qtokens, label, extra)
        # highest overlap first; highlights (deliberately marked cells) break ties above bulk rows; then a
        # stable order by source so the catalog is reproducible.
        return (-ov, 0 if rec.get("origin") == "highlight" else 1,
                str(rec.get("doc", "")), str(rec.get("sheet", "")), str(rec.get("cell", "")))

    raw.sort(key=_rank_key)
    out: list[OperandCandidate] = []
    for i, rec in enumerate(raw[:cap], 1):
        out.append(OperandCandidate(
            id=f"op{i}",
            value=rec.get("value"),
            number=float(rec.get("number", 0.0)),
            label=str(rec.get("label", "")),
            unit=str(rec.get("unit", "") or ""),
            source=str(rec.get("source") or _default_source(rec)),
            sheet=str(rec.get("sheet", "") or ""),
            cell=str(rec.get("cell", "") or ""),
            doc=str(rec.get("doc", "") or ""),
            origin=str(rec.get("origin", "")),
            context={k: rec[k] for k in ("group", "column", "color", "row_label")
                     if rec.get(k) not in (None, "")},
        ))
    return out


def _default_source(rec: dict[str, Any]) -> str:
    doc = str(rec.get("doc", "") or "")
    sheet = str(rec.get("sheet", "") or "")
    cell = str(rec.get("cell", "") or "")
    loc = f"{sheet}!{cell}" if sheet else cell
    return f"{doc}:{loc}" if doc and loc else (doc or loc)


def build_catalog(question: str, resolved_docs: Sequence[str], *,
                  project: str | None = None,
                  limit: int | None = None) -> list[dict[str, Any]]:
    """Enumerate candidates and render them as the JSON catalog for the Evidence Packet (may be empty)."""
    return [c.to_catalog_entry()
            for c in enumerate_candidates(question, resolved_docs, project=project, limit=limit)]


# --------------------------------------------------------------------------- directive rendering
def candidates_directive(catalog: Sequence[dict[str, Any]]) -> str:
    """The pre-inject block that turns operand discovery into a selection over the catalog.

    Lists each candidate as ``id / value(unit) / label / source`` and instructs the agent to *select*
    the operands by ``id`` (passing ``verify_formula`` a ``catalog`` + operands with ``{"name","select"}``)
    rather than re-typing values or grepping for them — value/source are taken verbatim from the catalog.
    Injects no answer: the catalog is candidate provenance, and which id is base/rate/条件 is the model's
    call, still checked by the PoT lane's operand/formula/execution layers.
    """
    if not catalog:
        return ""
    lines = []
    for c in catalog:
        val = c.get("value")
        unit = c.get("unit") or ""
        label = c.get("label") or ""
        src = c.get("source") or ""
        lines.append(f"  - {c.get('id')}: {val}{unit}"
                     + (f"（{label}）" if label else "")
                     + (f" [出典 {src}]" if src else ""))
    body = "\n".join(lines)
    return (
        "【operand 候補（決定論で事前列挙・SOT-2616）】この数値問いの operand は、対象文書から機械的に"
        "抽出した以下の候補から『選ぶ』だけにしてください（値を grep で探し直さない・書き写さない）:\n"
        f"{body}\n"
        "計算時は `verify_formula` に `catalog`（上の候補配列）を渡し、各 operand を "
        "{\"name\": <役割>, \"select\": <id>} で参照してください（value/unit/source は候補から自動束縛され"
        "ます）。候補に無い値が必要な場合のみ、従来どおり原文から根拠づけて追加してください。"
        "候補が問いに合致しないときは無理に選ばず、通常経路で最善の回答を返してください。")
