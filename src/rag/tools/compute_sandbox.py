"""compute_sandbox — deterministic pandas execution tool (暗算禁止・計算を証跡化).

An agent-callable tool that runs a *restricted* pandas expression against a corpus data
file and returns a uniform ``{value, evidence, method}`` contract, so every aggregate /
average / filter / ratio / 増減 is **computed and evidenced** rather than guessed by the LLM.

Design points that make this trustworthy on this specific corpus:

* **NFC path resolution.** The share drive is NFD-normalized (macOS origin); a plain
  ``open("…train.csv")`` fails on the Japanese folder names. File references resolve through
  :func:`src.rag.corpus.walk`, which is already NFC-aware.
* **``data_only`` xlsx reads.** ``train.xlsx`` here keeps its data on a ``train`` sheet while the
  *active* sheet is an empty ``グラフ`` chart — and formula cells only carry cached values. We load
  workbooks with ``data_only=True`` and auto-pick the real data sheet instead of the chart.
* **Safe restricted execution (allowed API only).** The expression is compiled in ``eval`` mode
  (statements are impossible) and AST-validated: only ``df`` / ``pd`` and a small set of pure
  builtins are reachable, dunder attribute traversal is blocked, and imports / lambdas /
  comprehensions are rejected — so there is no path from the expression to the interpreter, the
  filesystem, or the network.
"""
from __future__ import annotations

import ast
import math
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Mapping

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from src.rag.corpus import FileRef, nfc, walk


class ComputeError(ValueError):
    """Raised when a file cannot be resolved/read or an expression is rejected."""


# ---- allowed API (the only names an expression may reference) ----
_ALLOWED_NAMES = {"df", "pd"}
_SAFE_BUILTINS: dict[str, Any] = {
    "len": len, "abs": abs, "round": round, "min": min, "max": max, "sum": sum,
    "sorted": sorted, "float": float, "int": int, "str": str, "bool": bool,
    "list": list, "set": set, "dict": dict, "tuple": tuple, "range": range,
    "any": any, "all": all, "zip": zip, "enumerate": enumerate,
}
_READABLE_EXT = {"csv", "xlsx", "xlsm"}


# --------------------------------------------------------------------------- resolve
def _ref_from_path(path: Path) -> FileRef:
    p = Path(path)
    name = nfc(p.name)
    return FileRef(path=p, project="", category="", rel=name, name=name,
                   ext=p.suffix.lower().lstrip("."))


def _project_matches(ref: FileRef, hint: str) -> bool:
    """NFC substring match between a file's ``project`` and a caller-supplied project hint.

    Match is symmetric so a partial company name resolves either way: the question typically names
    the project loosely (e.g. ``青葉バイオメディカル機器``) while the corpus folder carries the full
    legal name (``株式会社青葉バイオメディカル機器``), and vice-versa. Empty projects never match a
    non-empty hint.
    """
    proj = nfc(ref.project or "")
    if not proj:
        return False
    return hint in proj or proj in hint


def _resolve(file: str | Path | FileRef, project: str | None = None) -> FileRef:
    """Resolve a file reference to an NFC-aware :class:`FileRef`.

    Accepts a :class:`FileRef`, an absolute/relative path, or a corpus-relative string matched
    (NFC) against the walked corpus by exact rel, rel-suffix, or filename.

    ``project`` optionally scopes an otherwise-ambiguous bare filename to a single case folder
    (source-location aid, SOT-2487): many projects each ship a ``train.xlsx`` / ``train.csv``, so
    ``compute("train.xlsx")`` alone is ambiguous, but ``project="青葉バイオメディカル機器"`` narrows
    it to that project's table. It only *narrows* candidates — it never widens or reorders them, so
    an already-unambiguous reference resolves identically whether or not a project is given.
    Ambiguous matches raise, and the error lists the candidate projects so the caller can retry with
    a ``project`` filter.
    """
    if isinstance(file, FileRef):
        return file
    if isinstance(file, Path) or (isinstance(file, str) and Path(file).is_absolute()):
        p = Path(file)
        want = nfc(str(p))
        for r in walk():
            if str(r.path) == str(p) or nfc(str(r.path)) == want:
                return r
        if p.exists():
            return _ref_from_path(p)
        raise ComputeError(f"file not found: {file}")

    key = nfc(str(file))
    refs = walk()
    hint = nfc(project) if project else None
    if hint:
        refs = [r for r in refs if _project_matches(r, hint)]
        if not refs:
            raise ComputeError(f"no corpus file matches {file!r} in project ~{project!r}")
    exact = [r for r in refs if nfc(r.rel) == key]
    if len(exact) == 1:
        return exact[0]
    suffix = [r for r in refs if nfc(r.rel).endswith(key)]
    if len(suffix) == 1:
        return suffix[0]
    named = [r for r in refs if nfc(r.name) == key]
    if len(named) == 1:
        return named[0]
    candidates = suffix or named or exact
    if not candidates:
        # last resort: treat as a filesystem path (only when unscoped)
        if hint is None:
            p = Path(file)
            if p.exists():
                return _ref_from_path(p)
        scope = f" in project ~{project!r}" if hint else ""
        raise ComputeError(f"no corpus file matches {file!r}{scope}")
    projects = sorted({r.project for r in candidates if r.project})
    hint_msg = (
        f"; disambiguate with project= (存在プロジェクト: {', '.join(projects[:8])})"
        if projects and hint is None else ""
    )
    raise ComputeError(
        f"ambiguous file reference {file!r} → {len(candidates)} matches: "
        + ", ".join(sorted(r.rel for r in candidates)[:5]) + hint_msg
    )


# --------------------------------------------------------------------------- read
def _coerce_numeric(df: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    """Best-effort per-column numeric coercion (parity with ``pd.read_csv`` dtype inference).

    Returns the coerced frame and the names of the columns that were converted, so the calculation
    trace can record exactly which normalization rules were applied to reach the input rows.
    """
    coerced: list[str] = []
    for col in df.columns:
        if df[col].dtype == object:
            try:
                converted = pd.to_numeric(df[col])
            except (ValueError, TypeError):
                continue
            df[col] = converted
            coerced.append(str(col))
    return df, coerced


def _pick_sheet(wb, sheet: int | str | None):
    if isinstance(sheet, int):
        return wb.worksheets[sheet]
    if isinstance(sheet, str):
        if sheet not in wb.sheetnames:
            raise ComputeError(f"sheet {sheet!r} not in {wb.sheetnames}")
        return wb[sheet]
    # auto: prefer a sheet literally named "train", else the largest data sheet, else active.
    for ws in wb.worksheets:
        if nfc(str(ws.title)).lower() == "train":
            return ws
    best, best_cells = None, -1
    for ws in wb.worksheets:
        rows, cols = ws.max_row or 0, ws.max_column or 0
        if rows >= 2 and cols >= 2 and rows * cols > best_cells:
            best, best_cells = ws, rows * cols
    return best or wb.active


def _read_xlsx(path: Path, sheet: int | str | None) -> tuple[pd.DataFrame, str, dict[str, Any]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = _pick_sheet(wb, sheet)
        title = str(ws.title)
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()
    # drop leading fully-empty rows (e.g. a title/blank line above the header)
    dropped_leading = 0
    while rows and all(c is None for c in rows[0]):
        rows.pop(0)
        dropped_leading += 1
    if not rows:
        raise ComputeError(f"sheet {title!r} in {path.name} has no data")
    header = rows[0]
    width = max((i + 1 for i, c in enumerate(header) if c is not None), default=len(header))
    columns = [nfc(str(c)) if c is not None else f"col_{i}" for i, c in enumerate(header[:width])]
    body = [(r[:width] + [None] * (width - len(r)))[:width] for r in rows[1:]]
    df, coerced = _coerce_numeric(pd.DataFrame(body, columns=columns))
    trace = _norm_trace(df, coerced, excluded_rows=dropped_leading,
                        extra_rules=([f"dropped_leading_empty_rows: {dropped_leading}"]
                                     if dropped_leading else []))
    return df, title, trace


def _norm_trace(df: pd.DataFrame, coerced: list[str], *, excluded_rows: int,
                extra_rules: list[str]) -> dict[str, Any]:
    """The structured calculation trace embedded in ``method.trace`` (SOT-2495 計算証跡の構造化出力).

    Records the rows the expression actually saw (``input_rows``), how many source rows were discarded
    while normalizing (``excluded_rows``), and the normalization rules applied (numeric coercion per
    column + any dropped rows) — the replayable input-shape证跡 the calc ledger stores per compute call.
    """
    rules = list(extra_rules)
    if coerced:
        rules.append("numeric_coercion: " + ", ".join(coerced))
    return {
        "input_rows": int(len(df)),
        "excluded_rows": int(excluded_rows),
        "normalization": rules,
    }


def _read_frame(ref: FileRef, sheet: int | str | None) -> tuple[pd.DataFrame, str | None, dict[str, Any]]:
    ext = ref.ext.lower()
    if ext not in _READABLE_EXT:
        raise ComputeError(f"unsupported file type .{ext} (only {sorted(_READABLE_EXT)})")
    if ext == "csv":
        df = pd.read_csv(ref.path)
        return df, None, _norm_trace(df, [], excluded_rows=0, extra_rules=[])
    return _read_xlsx(Path(ref.path), sheet)


# --------------------------------------------------------------------------- sandbox
def _validate(tree: ast.AST) -> None:
    for node in ast.walk(tree):
        if isinstance(node, (ast.Import, ast.ImportFrom)):
            raise ComputeError("imports are not allowed")
        if isinstance(node, ast.Lambda):
            raise ComputeError("lambda is not allowed")
        if isinstance(node, (ast.ListComp, ast.SetComp, ast.DictComp, ast.GeneratorExp)):
            raise ComputeError("comprehensions are not allowed")
        if isinstance(node, ast.Attribute) and node.attr.startswith("_"):
            raise ComputeError(f"access to private attribute {node.attr!r} is not allowed")
        if isinstance(node, ast.Name) and node.id not in _ALLOWED_NAMES and node.id not in _SAFE_BUILTINS:
            raise ComputeError(f"name {node.id!r} is not in the allowed API")


def _eval(expr: str, df: pd.DataFrame) -> Any:
    try:
        tree = ast.parse(expr, mode="eval")
    except SyntaxError as exc:
        raise ComputeError(f"invalid expression: {exc}") from exc
    _validate(tree)
    namespace = {"__builtins__": {}, "df": df, "pd": pd, **_SAFE_BUILTINS}
    return eval(compile(tree, "<compute_sandbox>", "eval"), namespace)  # noqa: S307 (sandboxed)


def _to_py(value: Any) -> Any:
    """Coerce a pandas/numpy result to a plain JSON-serializable Python value."""
    if isinstance(value, pd.Series):
        return [_to_py(v) for v in value.tolist()]
    if isinstance(value, pd.DataFrame):
        return [{str(k): _to_py(v) for k, v in row.items()} for row in value.to_dict("records")]
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    if hasattr(value, "item") and not isinstance(value, (str, bytes)):
        try:
            return value.item()
        except (ValueError, AttributeError):
            pass
    return value


# --------------------------------------------------------------------------- evidence
def _referenced_columns(tree: ast.AST, columns) -> list[str]:
    col_set = {str(c) for c in columns}
    hits: set[str] = set()
    for node in ast.walk(tree):
        if isinstance(node, ast.Constant) and isinstance(node.value, str) and node.value in col_set:
            hits.add(node.value)
        elif isinstance(node, ast.Attribute) and node.attr in col_set:
            hits.add(node.attr)
    return [str(c) for c in columns if str(c) in hits]


def _a1_range(df: pd.DataFrame, used: list[str]) -> str:
    n_rows = len(df)
    cols = list(map(str, df.columns))
    if not cols:
        return ""
    idxs = [cols.index(c) for c in used] if used else [0, len(cols) - 1]
    lo, hi = min(idxs), max(idxs)
    last_row = n_rows + 1  # +1 for the header row
    return f"{get_column_letter(lo + 1)}1:{get_column_letter(hi + 1)}{last_row}"


@dataclass
class _Ctx:
    ref: FileRef
    sheet: str | None


def run(
    file: str | Path | FileRef,
    expr: str,
    *,
    sheet: int | str | None = None,
    project: str | None = None,
    intermediates: Mapping[str, str] | None = None,
) -> dict[str, Any]:
    """Run ``expr`` against ``file`` and return the ``{value, evidence, method}`` contract.

    Parameters
    ----------
    file:
        A :class:`FileRef`, filesystem path, or corpus-relative string (NFC-resolved).
    expr:
        A single pandas expression referencing ``df`` (the loaded frame) and/or ``pd``.
    sheet:
        Optional xlsx sheet selector (index or name). Ignored for CSV. When omitted the real
        data sheet is auto-detected (a ``train`` sheet, else the largest, never an empty chart).
    project:
        Optional project/case-folder hint (NFC substring) that scopes an otherwise-ambiguous bare
        filename such as ``train.xlsx`` to a single project's table (source-location aid). Omit for
        a fully-qualified ``rel`` / path — behaviour is then identical to before.
    intermediates:
        Optional ``name -> sub-expression`` map; each is evaluated in the same sandbox and its
        value recorded under ``method.intermediates`` so the working is traceable.
    """
    ref = _resolve(file, project)
    df, sheet_used, trace = _read_frame(ref, sheet)

    tree = ast.parse(expr, mode="eval")  # (re-parsed inside _eval too; cheap and keeps _eval standalone)
    value = _to_py(_eval(expr, df))

    used = _referenced_columns(tree, df.columns)
    evidence = {
        "file": ref.rel or str(ref.path),
        "project": ref.project or None,
        "sheet": sheet_used,
        "rows": int(len(df)),
        "columns": [str(c) for c in df.columns],
        "columns_used": used,
        "range": _a1_range(df, used),
    }

    computed_intermediates: dict[str, Any] = {}
    for name, sub in (intermediates or {}).items():
        computed_intermediates[str(name)] = _to_py(_eval(sub, df))

    method = {
        "engine": "pandas",
        "code": expr,
        "allowed_api": sorted(_ALLOWED_NAMES | set(_SAFE_BUILTINS)),
        "intermediates": computed_intermediates,
        # SOT-2495: structured calculation trace (input rows / excluded rows / normalization rules) so a
        # numeric answer's derivation is replayable from the calc ledger. Additive — value is unchanged.
        "trace": trace,
    }
    return {"value": value, "evidence": evidence, "method": method}
