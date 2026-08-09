"""chart_numcache — read the *exact* plotted values of an embedded xlsx/pptx chart from its ``numCache``.

A great many corpus questions point at a **chart** ("グラフの◯◯の値は？", "系列Xの3月の数値は？").
A vision read of the rendered picture guesses those numbers from pixels and misreads decimals, axis
scale and overlapping bars. But an Office chart does not *store a picture* — it stores the data. Every
``.xlsx`` / ``.pptx`` chart is an OOXML ``c:chartSpace`` part (``xl/charts/chartN.xml`` /
``ppt/charts/chartN.xml``) whose each series carries a **cache of the values Excel last computed**:

* ``<c:val><c:numRef><c:numCache>`` — the plotted numbers (``<c:pt idx="i"><c:v>…</c:v></c:pt>``),
* ``<c:cat><c:strCache|c:numCache>`` — the category axis labels,
* ``<c:tx><c:strRef><c:strCache>`` (or a literal ``<c:v>``) — the series name,
* scatter/bubble charts use ``<c:xVal>`` / ``<c:yVal>`` / ``<c:bubbleSize>`` instead of ``cat``/``val``.

Because the cache is Excel's own snapshot of the source cells, ``numCache`` value ``==`` the cell's
computed value — so reading it is the *exact* plot value, not a vision estimate. This tool walks those
records deterministically and returns them under the unified :mod:`src.rag.tools.contract` shape
``{value, evidence, method}`` so the generation layer can call it exactly like the other tools.

* ``value``    — ``{charts, n_charts, n_series}``. ``charts`` is a list of ``{member, chart_types,
  series:[…]}``; each series is ``{name, categories, values, x_values, y_values, bubble_sizes,
  value_formula, cat_formula, format_code}`` (axis lists absent for that chart kind are ``[]``).
* ``evidence`` — ``file`` (+ ``member`` per chart), ``n_charts``, ``n_series`` and the source formulas.
* ``method``   — ``engine="chart_numcache"``; pure XML read (no secret, NFC-normalized text).

Accepts raw chart XML ``bytes``, a whole ``.xlsx`` / ``.pptx`` file (path / :class:`FileRef` / NFC
corpus-relative reference), and resolves NFD-on-disk names NFC-aware like the rest of the toolset.
"""
from __future__ import annotations

import hashlib
import io
import math
import re
import statistics
import zipfile
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from src.rag.corpus import FileRef, nfc
from src.rag.tools import contract
from src.rag.tools.contract import ContractError

# Chart parts live at ``<pkg>/charts/chartN.xml`` (xlsx: ``xl/charts/``, pptx: ``ppt/charts/``). The
# sibling ``colorsN.xml`` / ``styleN.xml`` are theme, not data — match only the numbered chart body.
_CHART_MEMBER = re.compile(r"charts/chart\d+\.xml$", re.IGNORECASE)

# The chart-type element local-names whose children hold ``c:ser`` (used only to report the kind).
_CHART_TYPE_TAGS = frozenset({
    "barChart", "bar3DChart", "lineChart", "line3DChart", "pieChart", "pie3DChart", "ofPieChart",
    "doughnutChart", "areaChart", "area3DChart", "scatterChart", "bubbleChart", "radarChart",
    "stockChart", "surfaceChart", "surface3DChart",
})


def _local(tag: str) -> str:
    """Local name of a namespaced ElementTree tag (``{ns}val`` → ``val``)."""
    return tag.rsplit("}", 1)[-1]


def _child(el: ET.Element, name: str) -> ET.Element | None:
    for c in el:
        if _local(c.tag) == name:
            return c
    return None


def _children(el: ET.Element, name: str):
    return [c for c in el if _local(c.tag) == name]


def _to_number(text: str) -> Any:
    """Parse a ``<c:v>`` numeric string to int when integral else float; keep the string if unparsable."""
    s = text.strip()
    try:
        f = float(s)
    except (TypeError, ValueError):
        return nfc(s)
    return int(f) if f.is_integer() and "." not in s and "e" not in s.lower() else f


def _read_points(axis: ET.Element) -> dict[str, Any]:
    """Decode a ``c:cat``/``c:val``/``c:tx``/``c:xVal``/``c:yVal``/``c:bubbleSize`` element.

    Returns ``{values, formula, format_code, numeric}`` where ``values`` is ordered by ``c:pt`` ``idx``
    (missing indices filled with ``None``). Handles the reference forms (``numRef``/``strRef``/
    ``multiLvlStrRef`` wrapping a ``numCache``/``strCache``), the inline literal forms (``numLit``/
    ``strLit``), and a bare ``<c:v>`` (a directly-typed series name). ``numeric`` marks a number cache
    so values are parsed to int/float (the exact plot value) rather than kept as label strings.
    """
    # A directly-typed value (common for series names: ``<c:tx><c:v>Name</c:v></c:tx>``).
    direct = _child(axis, "v")
    if direct is not None and direct.text is not None:
        return {"values": [nfc(direct.text)], "formula": None, "format_code": None, "numeric": False}

    ref = next((c for c in axis if _local(c.tag) in
                ("numRef", "strRef", "multiLvlStrRef", "numLit", "strLit")), None)
    if ref is None:
        return {"values": [], "formula": None, "format_code": None, "numeric": False}

    kind = _local(ref.tag)
    formula_el = _child(ref, "f")
    formula = nfc(formula_el.text) if formula_el is not None and formula_el.text else None

    # Where the ``pt`` list lives: a *Ref wraps a *Cache; a *Lit holds the pts itself.
    # (Use ``is not None`` — an empty ElementTree element is falsy, so ``or`` would skip a real cache.)
    if kind.endswith("Ref"):
        cache = None
        for cname in ("numCache", "strCache", "multiLvlStrCache"):
            cache = _child(ref, cname)
            if cache is not None:
                break
    else:
        cache = ref
    if cache is None:
        return {"values": [], "formula": formula, "format_code": None, "numeric": kind.startswith("num")}

    numeric = _local(cache.tag).startswith("num") or kind.startswith("num")
    fmt_el = _child(cache, "formatCode")
    format_code = nfc(fmt_el.text) if fmt_el is not None and fmt_el.text else None

    # multiLvlStrCache nests pts under c:lvl; flatten to the first (deepest-labelled) level's order.
    pt_parents = _children(cache, "lvl") or [cache]
    pts: dict[int, Any] = {}
    max_idx = -1
    for parent in pt_parents:
        for pt in _children(parent, "pt"):
            try:
                idx = int(pt.get("idx", "0"))
            except ValueError:
                continue
            v_el = _child(pt, "v")
            raw = v_el.text if v_el is not None else None
            val = _to_number(raw) if (numeric and raw is not None) else (nfc(raw) if raw is not None else None)
            pts.setdefault(idx, val)
            max_idx = max(max_idx, idx)

    count_el = _child(cache, "ptCount")
    try:
        n = int(count_el.get("val")) if count_el is not None else max_idx + 1
    except (TypeError, ValueError):
        n = max_idx + 1
    n = max(n, max_idx + 1)
    values = [pts.get(i) for i in range(n)]
    return {"values": values, "formula": formula, "format_code": format_code, "numeric": numeric}


def _series_name(ser: ET.Element) -> str | None:
    tx = _child(ser, "tx")
    if tx is None:
        return None
    vals = _read_points(tx)["values"]
    return vals[0] if vals else None


def _parse_series(ser: ET.Element) -> dict[str, Any]:
    """Turn one ``c:ser`` into ``{name, categories, values, x_values, y_values, bubble_sizes, …}``."""
    out: dict[str, Any] = {
        "name": _series_name(ser),
        "categories": [], "values": [], "x_values": [], "y_values": [], "bubble_sizes": [],
        "value_formula": None, "cat_formula": None, "format_code": None,
    }
    for tag, key, fkey in (
        ("cat", "categories", "cat_formula"),
        ("val", "values", "value_formula"),
        ("xVal", "x_values", None),
        ("yVal", "y_values", None),
        ("bubbleSize", "bubble_sizes", None),
    ):
        el = _child(ser, tag)
        if el is None:
            continue
        pts = _read_points(el)
        out[key] = pts["values"]
        if fkey:
            out[fkey] = pts["formula"]
        if tag == "val" and pts["format_code"]:
            out["format_code"] = pts["format_code"]
    return out


def _parse_chart_xml(data: bytes) -> dict[str, Any]:
    """Parse one ``c:chartSpace`` XML blob into ``{chart_types, series:[…]}``.

    Raises :class:`ContractError` if ``data`` is not a chart part (no ``c:ser`` / unparsable XML).
    """
    try:
        root = ET.fromstring(data)
    except ET.ParseError as e:
        raise ContractError(f"not a chart XML: {e}") from None

    chart_types: list[str] = []
    series: list[dict[str, Any]] = []
    for el in root.iter():
        ln = _local(el.tag)
        if ln in _CHART_TYPE_TAGS:
            chart_types.append(ln)
        elif ln == "ser":
            series.append(_parse_series(el))
    if not series:
        raise ContractError("chart XML has no <c:ser> series")
    return {"chart_types": sorted(set(chart_types)), "series": series}


# --------------------------------------------------------------------------- source loading
def _looks_like_zip(data: bytes) -> bool:
    return data[:4] in (b"PK\x03\x04", b"PK\x05\x06", b"PK\x07\x08")


def _read_source(file: str | Path | FileRef | bytes | bytearray) -> tuple[bytes, str, bool]:
    """Load bytes + a display label; ``is_office`` is True for an Office (zip) package, False for raw XML."""
    if isinstance(file, (bytes, bytearray)):
        data = bytes(file)
        return data, "<bytes>", _looks_like_zip(data)
    if isinstance(file, FileRef):
        return file.path.read_bytes(), file.rel or str(file.path), True
    p = Path(file)
    if p.exists():
        return p.read_bytes(), str(p), True
    from src.rag.tools.extract_tools import resolve_ref
    ref = resolve_ref(file)
    return ref.path.read_bytes(), ref.rel or str(ref.path), True


def chart_xml_members(office: str | Path | FileRef | bytes | bytearray) -> list[tuple[str, bytes]]:
    """Return ``(member_name, xml_bytes)`` for every ``charts/chartN.xml`` in an xlsx/pptx, sorted.

    Accepts the office file (path / :class:`FileRef` / NFC corpus-relative ref) or its raw zip bytes.
    """
    data, _label, is_office = _read_source(office)
    if not is_office:
        raise ContractError("not an Office package (expected .xlsx/.pptx zip bytes)")
    import io
    out: list[tuple[str, bytes]] = []
    with zipfile.ZipFile(io.BytesIO(data)) as zf:
        for name in sorted(zf.namelist()):
            if _CHART_MEMBER.search(name):
                out.append((name, zf.read(name)))
    return out


# --------------------------------------------------------------------------- public API
def _store_ref(file: str | Path | FileRef | bytes | bytearray) -> FileRef | None:
    """Resolve a file-like input to a corpus :class:`FileRef` for a structure-store lookup, else None.

    Only a real corpus file (with a ``rel``) can key the persisted store; raw chart-XML/zip bytes and
    unresolvable paths return None so the caller extracts live.
    """
    if isinstance(file, (bytes, bytearray)):
        return None
    try:
        from src.rag.tools.extract_tools import resolve_ref
        ref = file if isinstance(file, FileRef) else resolve_ref(file)
    except Exception:
        return None
    return ref if getattr(ref, "rel", None) else None


def extract_chart_numcache(file: str | Path | FileRef | bytes | bytearray) -> dict[str, Any]:
    """Read every chart's ``numCache``/``strCache`` from an xlsx/pptx (or raw chart XML) → contract.

    Parameters
    ----------
    file
        Raw chart-XML ``bytes``, a whole ``.xlsx``/``.pptx`` (as zip ``bytes``, a path, a
        :class:`FileRef`, or an NFC corpus-relative reference).

    Consults the deterministic structure pre-store (SOT-2533) first when enabled: a fresh cached
    contract is returned verbatim (byte-identical to a live read). Otherwise, and always for raw
    bytes, it reads live. ``value`` is ``{charts, n_charts, n_series}`` (see the module docstring).
    Raises :class:`ContractError` when the input holds no readable chart series.
    """
    ref = _store_ref(file)
    if ref is not None:
        from src.rag.index import structure_store
        cached = structure_store.lookup_chart_contract(ref)
        if cached is not None:
            return cached
    return _extract_chart_numcache_live(file)


def _extract_chart_numcache_live(file: str | Path | FileRef | bytes | bytearray) -> dict[str, Any]:
    """Live (store-bypassing) chart numCache read — the deterministic extraction core."""
    data, label, is_office = _read_source(file)

    parts: list[tuple[str, bytes]] = (
        chart_xml_members(data) if is_office else [(label, data)]
    )

    charts: list[dict[str, Any]] = []
    formulas: list[str] = []
    for member, blob in parts:
        try:
            parsed = _parse_chart_xml(blob)
        except ContractError:
            continue
        parsed["member"] = member if is_office else label
        charts.append(parsed)
        for s in parsed["series"]:
            for f in (s.get("value_formula"), s.get("cat_formula")):
                if f:
                    formulas.append(f)

    if not charts:
        raise ContractError(f"{label}: no readable chart numCache found")

    n_series = sum(len(c["series"]) for c in charts)
    value = {"charts": charts, "n_charts": len(charts), "n_series": n_series}
    return contract.make(
        value,
        engine="chart_numcache",
        evidence={
            "file": label,
            "n_charts": len(charts),
            "n_series": n_series,
            "members": [c["member"] for c in charts],
            "formulas": formulas,
        },
        nfc=True,
    )


def chart_series(file: str | Path | FileRef | bytes | bytearray) -> list[dict[str, Any]]:
    """Convenience: the flat list of every series across all charts (each tagged with its ``member``).

    Thin wrapper over :func:`extract_chart_numcache`; each series dict gains a ``member`` key so the
    caller can tell which chart it came from. Returns ``[]`` only if the file has charts but no series
    (``extract_chart_numcache`` raises when there is nothing at all).
    """
    result = extract_chart_numcache(file)
    out: list[dict[str, Any]] = []
    for chart in result["value"]["charts"]:
        for s in chart["series"]:
            out.append({**s, "member": chart["member"]})
    return out


def series_values(file: str | Path | FileRef | bytes | bytearray, *,
                  name: str | None = None) -> list[Any]:
    """Convenience: the exact plotted ``values`` of one series (by ``name``, else the first series).

    Matches ``name`` NFC-aware. Returns ``[]`` when no matching series exists. Useful for a direct
    "系列Xの数値は？" answer without the caller walking the nested contract.
    """
    want = nfc(name) if name is not None else None
    for s in chart_series(file):
        if want is None or (s.get("name") and nfc(str(s["name"])) == want):
            return s.get("values", [])
    return []


# --------------------------------------------------------------------------- raster-chart source fallback
# Excel's automatic histogram bin width follows Scott's normal reference rule.  The source workbooks
# persist the displayed interval width **truncated** (not rounded) to three decimal places, so the bin
# grid is ``min + k*width`` with that truncated width.  Using the unrounded NumPy ``bins=10`` default is
# a different chart and was the direct cause of SOT-2507 (1473 instead of the plotted 958).  Rounding the
# third decimal instead of truncating shifted the whole grid and mis-read the plotted bin (SOT-2546
# idx29: TP raw width 0.2008710 rounds to 0.201 → bin (6.102138, 6.303138] but the plotted width is the
# truncated 0.200 → [6.088138, 6.288138]); truncation still reproduces AG_ratio's 0.053 / 958 (idx10).
_ID_HEADERS = frozenset({"id", "index", "row_id", "record_id"})


def _image_anchor(image: Any) -> tuple[int, int]:
    """Return an embedded image's zero-based ``(row, column)`` anchor, fail-closed when unavailable."""
    anchor = getattr(image, "anchor", None)
    marker = getattr(anchor, "_from", None)
    if marker is None:
        raise ContractError("embedded chart image has no worksheet anchor")
    return int(marker.row), int(marker.col)


def _header_row(ws: Any) -> tuple[int, list[Any]] | None:
    """Find a table header near the top of a worksheet.

    A source table must have at least two non-empty string headers.  Ranking by header coverage and
    data-row count makes the real ``train`` sheet win over auxiliary pivot/output sheets without using
    a corpus-specific sheet name.
    """
    best: tuple[int, int, list[Any]] | None = None
    for row_idx in range(1, min(int(ws.max_row), 20) + 1):
        values = [ws.cell(row_idx, c).value for c in range(1, int(ws.max_column) + 1)]
        strings = sum(isinstance(v, str) and bool(v.strip()) for v in values)
        nonempty = sum(v is not None and bool(str(v).strip()) for v in values)
        if strings < 2:
            continue
        score = nonempty * 10_000 + min(max(int(ws.max_row) - row_idx, 0), 9_999)
        if best is None or score > best[0]:
            best = (score, row_idx, values)
    return (best[1], best[2]) if best is not None else None


def _source_columns(ws: Any, header_row: int, headers: list[Any]) -> list[tuple[int, str]]:
    """Numeric columns represented by a histogram grid, preserving source-table order.

    Chart sheets commonly include a numeric target (for example a binary 0/1 outcome) but omit text
    features, so target-name heuristics are unsafe.  Inspecting the cells makes the mapping portable.
    """
    out: list[tuple[int, str]] = []
    for col_idx, raw in enumerate(headers, 1):
        if raw is None or not str(raw).strip():
            continue
        name = nfc(str(raw).strip())
        folded = name.casefold()
        if folded in _ID_HEADERS:
            continue
        sample = [ws.cell(row, col_idx).value
                  for row in range(header_row + 1, min(int(ws.max_row), header_row + 200) + 1)]
        present = [v for v in sample if v is not None and str(v).strip()]
        numeric = [v for v in present
                   if isinstance(v, (int, float)) and not isinstance(v, bool)
                   and math.isfinite(float(v))]
        if not present or len(numeric) / len(present) < 0.8:
            continue
        out.append((col_idx, name))
    return out


def embedded_chart_sources(file: str | Path | FileRef | bytes | bytearray) -> list[dict[str, Any]]:
    """Map raster charts in an xlsx chart sheet to their authoritative source-table columns.

    Some workbooks flatten every histogram to PNG, leaving no ``c:numCache`` at all.  Their chart grid
    and source table retain the same feature order.  This function resolves that relationship only when
    it is unambiguous: one ordered image per non-ID numeric source column.  A mismatch raises instead
    of guessing.  Each mapping includes the image's decoded-byte SHA-256 so the chart artifact is pinned
    and auditable, but pixel text is never used as numeric evidence.
    """
    data, label, is_office = _read_source(file)
    if not is_office:
        raise ContractError("raster chart source mapping requires an xlsx Office package")

    try:
        wb = load_workbook(io.BytesIO(data), data_only=True, read_only=False)
    except Exception as exc:  # noqa: BLE001 - normalize parser failures to the tool contract
        raise ContractError(f"{label}: cannot open workbook for chart source mapping: {exc}") from None

    images: list[tuple[int, int, str, str, str]] = []
    with zipfile.ZipFile(io.BytesIO(data)) as zf:
        media_by_hash = {
            hashlib.sha256(zf.read(member)).hexdigest(): member
            for member in zf.namelist() if member.startswith("xl/media/")
        }
    for ws in wb.worksheets:
        for image in getattr(ws, "_images", ()):
            row, col = _image_anchor(image)
            blob = image._data()  # openpyxl exposes the original embedded bytes; no OCR/vision involved
            digest = hashlib.sha256(blob).hexdigest()
            images.append((row, col, ws.title, media_by_hash.get(digest, "<embedded-image>"), digest))
    if not images:
        raise ContractError(f"{label}: no embedded chart images found")
    images.sort(key=lambda item: (item[0], item[1]))

    candidates: list[tuple[int, Any, int, list[Any]]] = []
    image_sheets = {item[2] for item in images}
    for ws in wb.worksheets:
        if ws.title in image_sheets:
            continue
        header = _header_row(ws)
        if header is None:
            continue
        row_idx, headers = header
        score = len(_source_columns(ws, row_idx, headers)) * 10_000 + min(int(ws.max_row), 9_999)
        candidates.append((score, ws, row_idx, headers))
    if not candidates:
        raise ContractError(f"{label}: no source data table found for embedded charts")
    _score, source_ws, header_row, headers = max(candidates, key=lambda item: item[0])
    columns = _source_columns(source_ws, header_row, headers)
    if len(columns) != len(images):
        raise ContractError(
            f"{label}: ambiguous image/source mapping ({len(images)} images, {len(columns)} feature columns)")

    last_row = int(source_ws.max_row)
    out: list[dict[str, Any]] = []
    for (row, col, chart_sheet, member, digest), (source_col, column) in zip(images, columns):
        letter = get_column_letter(source_col)
        out.append({
            "member": member,
            "pixel_sha256": digest,
            "chart_sheet": nfc(chart_sheet),
            "anchor": {"row": row, "column": col},
            "source_sheet": nfc(source_ws.title),
            "source_column": column,
            "source_range": f"{nfc(source_ws.title)}!${letter}${header_row + 1}:${letter}${last_row}",
            "source_column_index": source_col,
        })
    return out


def _scott_histogram(values: list[float]) -> tuple[list[int], list[str], float]:
    """Recompute an Excel-style automatic histogram from source values (Scott width, no pixels)."""
    clean = [float(v) for v in values if isinstance(v, (int, float)) and not isinstance(v, bool)
             and math.isfinite(float(v))]
    if len(clean) < 2 or min(clean) == max(clean):
        raise ContractError("histogram source column needs at least two distinct numeric values")
    raw_width = 3.5 * statistics.pstdev(clean) * (len(clean) ** (-1.0 / 3.0))
    # The workbook chart generator stores automatic widths *truncated* to three decimal places (Excel
    # ROUNDDOWN semantics), not rounded — rounding shifts the whole bin grid off by one (SOT-2546 idx29).
    # The 1e-9 nudge guards against a float representation landing just under an exact third-decimal.
    # If a very small scale would truncate to zero, retain three significant digits instead.
    width = math.floor(raw_width * 1000.0 + 1e-9) / 1000.0
    if width <= 0:
        width = float(f"{raw_width:.3g}")
    start = min(clean)
    n_bins = max(1, math.ceil((max(clean) - start) / width))
    counts = [0] * n_bins
    for value in clean:
        idx = min(int((value - start) / width), n_bins - 1)
        counts[idx] += 1
    categories = []
    for idx in range(n_bins):
        left = start + idx * width
        right = left + width
        categories.append(f"{'[' if idx == 0 else '('}{left:.12g}, {right:.12g}]")
    return counts, categories, width


def read_chart_values(file: str | Path | FileRef | bytes | bytearray, *,
                      column: str | None = None,
                      operation: str | None = None) -> dict[str, Any]:
    """Strict numeric chart path: numCache first, otherwise source-table recomputation.

    ``vision`` is intentionally absent.  Raster charts are numeric-authoritative only when their image
    maps unambiguously to a source column and the requested statistic can be recomputed from those cells.
    ``operation='histogram_max_count'`` returns the maximum bin count as ``value.result`` while retaining
    the full bin ledger.  Unsupported/ambiguous cases raise so callers abstain rather than commit pixels.
    """
    try:
        exact = extract_chart_numcache(file)
    except ContractError:
        exact = None
    if exact is not None:
        exact["method"]["numeric_authority"] = True
        exact["method"]["vision_used"] = False
        return exact

    if not column:
        raise ContractError("embedded-image chart requires source column for deterministic recomputation")
    op = operation or "histogram"
    if op not in {"histogram", "histogram_max_count"}:
        raise ContractError(f"unsupported raster chart operation: {op}")
    mappings = embedded_chart_sources(file)
    want = nfc(column).casefold()
    matches = [m for m in mappings if m["source_column"].casefold() == want]
    if len(matches) != 1:
        raise ContractError(f"source column not uniquely mapped to an embedded chart: {column}")
    mapping = matches[0]

    data, label, _is_office = _read_source(file)
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=False)
    ws = wb[mapping["source_sheet"]]
    source_col = int(mapping["source_column_index"])
    values = [ws.cell(row, source_col).value for row in range(2, int(ws.max_row) + 1)]
    counts, categories, width = _scott_histogram(values)
    series = {
        "name": mapping["source_column"],
        "categories": categories,
        "values": counts,
        "x_values": [], "y_values": [], "bubble_sizes": [],
        "value_formula": mapping["source_range"],
        "cat_formula": None,
        "format_code": "count",
        "source_sheet": mapping["source_sheet"],
        "source_range": mapping["source_range"],
        "bin_width": width,
        "derivation": "Scott normal reference rule over source cells",
    }
    result = max(counts) if op == "histogram_max_count" else counts
    value = {
        "charts": [{
            "member": mapping["member"],
            "chart_types": ["imageHistogram"],
            "pixel_sha256": mapping["pixel_sha256"],
            "series": [series],
        }],
        "n_charts": 1,
        "n_series": 1,
        "operation": op,
        "result": result,
    }
    return contract.make(
        value,
        engine="chart_source_compute",
        evidence={
            "file": label,
            "member": mapping["member"],
            "pixel_sha256": mapping["pixel_sha256"],
            "source_sheet": mapping["source_sheet"],
            "source_column": mapping["source_column"],
            "source_range": mapping["source_range"],
            "n_rows": sum(counts),
            "bin_width": width,
        },
        nfc=True,
        numeric_authority=True,
        vision_used=False,
    )
