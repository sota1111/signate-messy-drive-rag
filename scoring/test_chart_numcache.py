"""SOT-2462: exact xlsx/pptx chart plot-value extraction from ``numCache`` (``chart_numcache``).

Covers the two acceptance criteria:
  ① a chart question returns the **exact** plotted value — series name, categories and the numeric
     array are read straight from ``<c:numCache>``/``<c:strCache>`` (incl. scatter ``xVal``/``yVal``),
     and the cached numbers equal the source cells' computed values;
  ② every result is the common contract ``{value, evidence, method}``.

Behaviour is pinned against tiny, hand-built chart XML whose records we control exactly (fast, no heavy
fixtures, no committing non-redistributable SIGNATE data). The "numCache == cell computed value"
invariant is proved end-to-end by building a real ``.xlsx`` (data cells via openpyxl) and injecting a
chart part whose ``numCache`` mirrors those cells, then asserting the tool reads back exactly the
openpyxl cell values. A guarded smoke test exercises the real share drive.
"""
from __future__ import annotations

import io
import zipfile

import openpyxl
import pytest

from src.rag.tools import (
    ContractError,
    chart_series,
    chart_xml_members,
    embedded_chart_sources,
    extract_chart_numcache,
    is_contract,
    read_chart_values,
    series_values,
)

C_NS = "http://schemas.openxmlformats.org/drawingml/2006/chart"


# --------------------------------------------------------------------------- minimal chart XML builder
def _pts(values, numeric: bool) -> str:
    tag = "c:v"
    out = [f'<c:ptCount val="{len(values)}"/>']
    for i, v in enumerate(values):
        if v is None:
            continue
        out.append(f'<c:pt idx="{i}"><{tag}>{v}</{tag}></c:pt>')
    return "".join(out)


def _num_ref(formula: str, values, fmt: str = "General") -> str:
    return (f"<c:numRef><c:f>{formula}</c:f><c:numCache>"
            f"<c:formatCode>{fmt}</c:formatCode>{_pts(values, True)}</c:numCache></c:numRef>")


def _str_ref(formula: str, values) -> str:
    return (f"<c:strRef><c:f>{formula}</c:f><c:strCache>"
            f"{_pts(values, False)}</c:strCache></c:strRef>")


def _ser(idx: int, name: str, name_ref: str, cats, cat_ref: str, vals, val_ref: str) -> str:
    return (f'<c:ser><c:idx val="{idx}"/><c:order val="{idx}"/>'
            f"<c:tx>{_str_ref(name_ref, [name])}</c:tx>"
            f"<c:cat>{_str_ref(cat_ref, cats)}</c:cat>"
            f"<c:val>{_num_ref(val_ref, vals)}</c:val></c:ser>")


def _bar_chart(*sers: str) -> bytes:
    body = "".join(sers)
    xml = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
           f'<c:chartSpace xmlns:c="{C_NS}"><c:chart><c:plotArea>'
           f'<c:barChart><c:barDir val="col"/><c:grouping val="clustered"/>{body}</c:barChart>'
           "</c:plotArea></c:chart></c:chartSpace>")
    return xml.encode("utf-8")


# A 2-series column chart: "売上"(sales) and "原価"(cost) over 3 months, with a gap (idx 3 missing).
_SALES = [1.5, 3.25, 7.0]
_COST = [1.0, 2.5, 4.75]
_MONTHS = ["1月", "2月", "3月"]


@pytest.fixture()
def bar_xml() -> bytes:
    return _bar_chart(
        _ser(0, "売上", "Sheet1!$B$1", _MONTHS, "Sheet1!$A$2:$A$4", _SALES, "Sheet1!$B$2:$B$4"),
        _ser(1, "原価", "Sheet1!$C$1", _MONTHS, "Sheet1!$A$2:$A$4", _COST, "Sheet1!$C$2:$C$4"),
    )


# --------------------------------------------------------------------------- contract (②)
def test_returns_contract(bar_xml: bytes):
    out = extract_chart_numcache(bar_xml)
    assert is_contract(out)
    assert out["method"]["engine"] == "chart_numcache"
    v = out["value"]
    assert set(v) == {"charts", "n_charts", "n_series"}
    assert v["n_charts"] == 1 and v["n_series"] == 2
    assert out["evidence"]["formulas"]  # source cell ranges recorded


# --------------------------------------------------------------------------- exact values (①)
def test_exact_series_values_and_names(bar_xml: bytes):
    series = chart_series(bar_xml)
    assert [s["name"] for s in series] == ["売上", "原価"]
    assert series[0]["values"] == _SALES          # exact floats, not vision estimates
    assert series[1]["values"] == _COST
    assert series[0]["categories"] == _MONTHS
    assert series[0]["value_formula"] == "Sheet1!$B$2:$B$4"


def test_series_values_by_name(bar_xml: bytes):
    assert series_values(bar_xml, name="原価") == _COST
    assert series_values(bar_xml) == _SALES        # first series when name omitted
    assert series_values(bar_xml, name="存在しない") == []


def test_chart_type_reported(bar_xml: bytes):
    chart = extract_chart_numcache(bar_xml)["value"]["charts"][0]
    assert chart["chart_types"] == ["barChart"]


def test_integer_cache_stays_integer():
    xml = _bar_chart(_ser(0, "n", "S!$B$1", ["x"], "S!$A$2", [42], "S!$B$2"))
    vals = chart_series(xml)[0]["values"]
    assert vals == [42] and isinstance(vals[0], int)


def test_missing_point_is_gap():
    # ptCount=3 but idx 1 absent → the hole is preserved as None (not silently shifted).
    xml = (f'<c:chartSpace xmlns:c="{C_NS}"><c:chart><c:plotArea><c:lineChart><c:ser>'
           '<c:idx val="0"/><c:tx><c:v>g</c:v></c:tx>'
           '<c:val><c:numRef><c:f>S!$B$2:$B$4</c:f><c:numCache><c:ptCount val="3"/>'
           '<c:pt idx="0"><c:v>10</c:v></c:pt><c:pt idx="2"><c:v>30</c:v></c:pt>'
           "</c:numCache></c:numRef></c:val></c:ser></c:lineChart></c:plotArea></c:chart></c:chartSpace>")
    s = chart_series(xml.encode("utf-8"))[0]
    assert s["name"] == "g"          # series name from a direct <c:tx><c:v>
    assert s["values"] == [10, None, 30]


def test_scatter_x_and_y_values():
    xml = (f'<c:chartSpace xmlns:c="{C_NS}"><c:chart><c:plotArea><c:scatterChart><c:ser>'
           '<c:idx val="0"/><c:tx><c:strRef><c:f>S!$B$1</c:f>'
           '<c:strCache><c:ptCount val="1"/><c:pt idx="0"><c:v>点</c:v></c:pt></c:strCache></c:strRef></c:tx>'
           '<c:xVal><c:numRef><c:f>S!$A$2:$A$3</c:f><c:numCache><c:ptCount val="2"/>'
           '<c:pt idx="0"><c:v>0.1</c:v></c:pt><c:pt idx="1"><c:v>0.2</c:v></c:pt></c:numCache></c:numRef></c:xVal>'
           '<c:yVal><c:numRef><c:f>S!$B$2:$B$3</c:f><c:numCache><c:ptCount val="2"/>'
           '<c:pt idx="0"><c:v>9.9</c:v></c:pt><c:pt idx="1"><c:v>8.8</c:v></c:pt></c:numCache></c:numRef></c:yVal>'
           "</c:ser></c:scatterChart></c:plotArea></c:chart></c:chartSpace>")
    s = chart_series(xml.encode("utf-8"))[0]
    assert s["x_values"] == [0.1, 0.2]
    assert s["y_values"] == [9.9, 8.8]
    assert s["values"] == []          # scatter has no c:val axis


# --------------------------------------------------------------------------- numCache == cell value (①)
def _xlsx_with_chart_numcache() -> tuple[bytes, list[float], list[float]]:
    """A real .xlsx: a data sheet with literal cells + a chart part whose numCache mirrors those cells.

    openpyxl writes a chart with only ``<c:f>`` (no cache), so we replace the chart part with one whose
    ``numCache`` carries the exact cell values — reproducing what Excel itself embeds. Returns the file
    bytes and the two expected value columns.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    for row in ([None, "売上", "原価"],
                ["1月", _SALES[0], _COST[0]],
                ["2月", _SALES[1], _COST[1]],
                ["3月", _SALES[2], _COST[2]]):
        ws.append(row)
    from openpyxl.chart import BarChart, Reference
    ch = BarChart()
    ch.add_data(Reference(ws, min_col=2, max_col=3, min_row=1, max_row=4), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=1, min_row=2, max_row=4))
    ws.add_chart(ch, "E1")
    buf = io.BytesIO()
    wb.save(buf)

    chart_xml = _bar_chart(
        _ser(0, "売上", "Data!$B$1", _MONTHS, "Data!$A$2:$A$4", _SALES, "Data!$B$2:$B$4"),
        _ser(1, "原価", "Data!$C$1", _MONTHS, "Data!$A$2:$A$4", _COST, "Data!$C$2:$C$4"),
    )
    src = zipfile.ZipFile(buf)
    out_buf = io.BytesIO()
    with zipfile.ZipFile(out_buf, "w", zipfile.ZIP_DEFLATED) as dst:
        for item in src.infolist():
            data = src.read(item.filename)
            if item.filename.endswith("charts/chart1.xml"):
                data = chart_xml
            dst.writestr(item, data)
    return out_buf.getvalue(), _SALES, _COST


def test_numcache_matches_cell_computed_values():
    xlsx_bytes, sales, cost = _xlsx_with_chart_numcache()

    # source of truth: the cells' computed values, read independently via openpyxl.
    ws = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)["Data"]
    cell_sales = [ws.cell(row=r, column=2).value for r in (2, 3, 4)]
    cell_cost = [ws.cell(row=r, column=3).value for r in (2, 3, 4)]
    assert cell_sales == sales and cell_cost == cost

    # the tool reads the chart numCache and it equals those cell values (exact, not vision-estimated).
    series = chart_series(xlsx_bytes)
    by_name = {s["name"]: s["values"] for s in series}
    assert by_name["売上"] == cell_sales
    assert by_name["原価"] == cell_cost


def test_chart_xml_members_finds_chart_in_xlsx():
    xlsx_bytes, _s, _c = _xlsx_with_chart_numcache()
    members = chart_xml_members(xlsx_bytes)
    assert members and all(name.endswith(".xml") and "chart" in name for name, _ in members)


# --------------------------------------------------------------------------- raster chart -> source-table fallback (SOT-2507)
def _xlsx_with_raster_histograms() -> bytes:
    """A chart sheet whose PNG order mirrors numeric source columns (text columns are not charted)."""
    from PIL import Image as PILImage
    from openpyxl.drawing.image import Image as XLImage

    wb = openpyxl.Workbook()
    charts = wb.active
    charts.title = "Charts"
    source = wb.create_sheet("source")
    source.append(["id", "text_feature", "metric", "target"])
    for idx, value in enumerate([0.1, 0.2, 0.21, 0.22, 0.4, 0.41], 1):
        source.append([idx, f"row-{idx}", value, idx % 2])
    # Numeric source order excluding id is metric -> target; anchor order is A1 -> H1.
    streams = []
    for color, anchor in [((20, 40, 60), "A1"), ((80, 100, 120), "H1")]:
        stream = io.BytesIO()
        PILImage.new("RGB", (24, 16), color).save(stream, format="PNG")
        stream.seek(0)
        streams.append(stream)  # keep alive until the workbook is serialized
        charts.add_image(XLImage(stream), anchor)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def test_embedded_chart_sources_maps_anchor_order_to_numeric_columns():
    xlsx = _xlsx_with_raster_histograms()
    mapped = embedded_chart_sources(xlsx)
    assert [m["source_column"] for m in mapped] == ["metric", "target"]
    assert mapped[0]["source_range"] == "source!$C$2:$C$7"
    assert mapped[0]["pixel_sha256"] != mapped[1]["pixel_sha256"]


def test_raster_histogram_is_recomputed_from_source_without_vision():
    out = read_chart_values(
        _xlsx_with_raster_histograms(), column="metric", operation="histogram_max_count")
    assert is_contract(out)
    assert out["method"] == {
        "engine": "chart_source_compute", "nfc": True,
        "numeric_authority": True, "vision_used": False,
    }
    assert out["value"]["result"] == max(out["value"]["charts"][0]["series"][0]["values"])
    assert out["evidence"]["source_column"] == "metric"


def test_raster_chart_without_column_fails_closed():
    with pytest.raises(ContractError, match="requires source column"):
        read_chart_values(_xlsx_with_raster_histograms())


def test_unified_reader_keeps_numcache_as_first_authority(bar_xml: bytes):
    out = read_chart_values(bar_xml)
    assert out["method"]["engine"] == "chart_numcache"
    assert out["method"]["numeric_authority"] is True
    assert out["method"]["vision_used"] is False


# --------------------------------------------------------------------------- input guards
def test_non_chart_bytes_raise():
    with pytest.raises(ContractError):
        extract_chart_numcache(b"<not-a-chart/>")


def test_office_without_charts_raises():
    wb = openpyxl.Workbook()
    wb.active["A1"] = 1
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(ContractError):
        extract_chart_numcache(buf.getvalue())


# --------------------------------------------------------------------------- real corpus smoke
def test_real_corpus_chart_numcache_smoke():
    from src.rag.corpus import walk
    ref = next((r for r in walk() if r.ext in ("xlsx", "pptx")
                and chart_xml_members_safe(r)), None)
    if ref is None:
        pytest.skip("no corpus file with an embedded chart present")
    out = extract_chart_numcache(ref)
    assert is_contract(out)
    assert out["value"]["n_series"] >= 1
    # at least one series exposes a numeric plot value.
    assert any(isinstance(v, (int, float))
               for c in out["value"]["charts"] for s in c["series"] for v in s["values"])


def test_real_corpus_idx10_raster_histogram_recomputes_958():
    """Gold idx10: source-cell Scott bins reproduce the plotted maximum, not NumPy bins=10 (1473)."""
    from src.rag.corpus import walk
    ref = next((r for r in walk() if r.ext == "xlsx" and r.path.name == "train.xlsx"
                and "かえで総合病院" in r.rel), None)
    if ref is None:
        pytest.skip("idx10 corpus workbook not present")
    out = read_chart_values(ref, column="AG_ratio", operation="histogram_max_count")
    assert out["value"]["result"] == 958
    assert out["evidence"]["source_range"] == "train!$K$2:$K$3501"
    assert out["method"]["vision_used"] is False


def chart_xml_members_safe(ref) -> bool:
    try:
        return bool(chart_xml_members(ref))
    except Exception:
        return False
